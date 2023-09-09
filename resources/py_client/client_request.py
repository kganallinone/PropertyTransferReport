import tkinter as tk
from tkinter import ttk
import customtkinter as ctk
from PIL import Image
from tkcalendar import *
from tkinter import messagebox
import pyrebase
from pathlib import Path
from openpyxl import load_workbook
import openpyxl
from openpyxl import Workbook
import datetime
import os

ctk.set_appearance_mode("light")  # Modes: system (default), light, dark
current_datetime = datetime.datetime.now()
current_day = current_datetime.strftime("%Y-%m-%d")
current_time = current_datetime.strftime("%H:%M:%S")

def request_window():
    #from resources.py_client.client_dashboard import dashboard_window_client
    global suggestions_frame
    suggestions_frame = None
    suggestions_frame = None
    report_form = ctk.CTkToplevel()
    report_form.title("Property Transfer | Report")
    report_form.geometry("{0}x{1}+200+25".format(1000,650))
    report_form.resizable(False, False)
    report_form.configure(fg_color = 'white')

    #<----------------FIREBASE CONFIGURATION--------------------------------->#
    config = {
        "apiKey": "AIzaSyCyCtXg05ff-tocdUaxjBFfa550TLfKZQ4",
        "authDomain": "puplopez-ptp.firebaseapp.com",
        "projectId": "puplopez-ptp",
        "databaseURL": "https://puplopez-ptp-default-rtdb.firebaseio.com/",
        "storageBucket": "puplopez-ptp.appspot.com",
        "messagingSenderId": "638312293451",
        "appId": "1:638312293451:web:79e9405d11db0496f599ce",
        "measurementId": "G-9Z3SPKFR9V"
    }
    #<----------------FIREBASE CONNECT--------------------------------->#
    firebase = pyrebase.initialize_app(config)

    db = firebase.database()
    
    wb = load_workbook("./resources/user/data/UserLog.xlsx", data_only=True)
    sh = wb["User email"]
    email = sh["A1"].value
    print(email)
    
    # Load the workbook
    workbook_refresh = openpyxl.load_workbook('./resources/user/data/items.xlsx')
    # Select the active worksheet
    worksheet_refresh = workbook_refresh.active
    # Iterate over each cell and set its value to None
    for row in worksheet_refresh.iter_rows():
        for cell in row:
            cell.value = None
    # Save the workbook
    workbook_refresh.save('./resources/user/data/items.xlsx')

    #<-------------------------ITEM SUGGESTION------------------------------------>
    def show_data(item):
        selected_item.set(item)
        clear_suggestions_frame()
        try:
            item_db = db.child("items_info").order_by_child("NO").equal_to(item).get()
            
            for dt in item_db.each():
                var1 = dt.val()['NO']
                var2 = dt.val()['NAME']
                var3 = dt.val()['DES']
                var4 = dt.val()['LOC']
                var5 = dt.val()['QTY']
                var6 = dt.val()['CON']
                print(var6)
                
            
            result_property_no.delete(0,len(var1))
            result_property_no.insert(0,var1)
            result_name.delete(0,len(var2))
            result_name.insert(0,var2)
            result_desc.delete(0,len(var3))
            result_desc.insert(0,var3)
            location_entry.delete(0,len(var4))
            location_entry.insert(0,var4)
            condition.set(var6)
            qty_var.set(var5)
            
            
            
        except:
            print(0) 

    def search_items(event):
        search_term = property_no.get().lower()
        if len(search_term) == 0:
            clear_suggestions_frame()
            return

        # Retrieve items data from Firebase
        snapshot = db.child("items_info").get()
        items = [item.key() for item in snapshot.each() if search_term in str(item.val()).lower()]

        # Clear previous suggestions
        clear_suggestions_frame()

        if items:
            create_suggestions_frame()

            for i, item in enumerate(items):
                button = ctk.CTkButton(suggestions_frame, text=item, width=property_no.winfo_width(), command=lambda i=item: show_data(i))
                button.pack(pady=1)
                if i == 9:  # Limit the number of suggestions displayed
                    break

            suggestions_frame.lift()

    def create_suggestions_frame():
        global suggestions_frame
        if suggestions_frame is None:
            suggestions_frame = tk.Frame(report_form)
            suggestions_frame.configure(bd=1, relief=tk.FLAT, bg="white", highlightbackground='white')
            suggestions_frame.configure(width=property_no.winfo_width())
            suggestions_frame.place(in_=sub_frame4_3_2, x=0, rely=1, relwidth=1, y=1)
            
        else:
            suggestions_frame.configure(width=property_no.winfo_width())
            
    def clear_suggestions_frame():
        global suggestions_frame
        if suggestions_frame is not None:
            suggestions_frame.destroy()
            suggestions_frame = None
    #<-------------------------ITEM SUGGESTION END------------------------------------>
     
    #<-------------------------LOCATION SUGGESTION------------------------------------>
    global suggestion_location_frame
    suggestion_location_frame = None

    def show_data2(location):
        selected_location.set(location)
        clear_suggestion_location_frame()
        location_entry.delete(0, tk.END)  # Clear the entry field
        location_entry.insert(0, location)  

    def search_locations(event):
        search_term = location_entry.get().lower()
        if len(search_term) == 0:
            clear_suggestion_location_frame()
            return

        # Replace data retrieval from Firebase with a string
        locations = locations = [
            "Computer Laboratory 1", 
            "Computer Laboratory 2", 
            "Computer Laboratory 3"
            ]

        locations = [location for location in locations if search_term in location.lower()]

        # Rest of the code remains the same...
        clear_suggestion_location_frame()

        if locations:
            create_suggestion_location_frame()

            for i, location in enumerate(locations):
                button = ttk.Button(suggestion_location_frame, text=location, width=location_entry.winfo_width(), command=lambda i=location: show_data2(i))
                button.pack(pady=1)
                if i == 4:  # Limit the number of suggestions displayed
                    break

            suggestion_location_frame.lift()

    def create_suggestion_location_frame():
        global suggestion_location_frame
        if suggestion_location_frame is None:
            suggestion_location_frame = tk.Frame(report_form)
            suggestion_location_frame.configure(bd=1, relief=tk.FLAT, bg="white")
            suggestion_location_frame.place(in_=frame_loc, x=0, rely=1, relwidth=1, y=1)
            location_entry.bind("<Configure>", update_suggestion_location_frame_width)  # Bind the Configure event
        else:
            suggestion_location_frame.configure(width=location_entry.winfo_width())

    def update_suggestion_location_frame_width(event):
        suggestion_location_frame.configure(width=location_entry.winfo_width())

    def clear_suggestion_location_frame():
        global suggestion_location_frame
        if suggestion_location_frame is not None:
            suggestion_location_frame.destroy()
            suggestion_location_frame = None

            
    #<-------------------------LOCATION SUGGESTION END------------------------------------>        
    def dashboard_open():
        report_form.withdraw()
        #dashboard_window_client()
    
    def add_entry():
        add_ptrno = result_property_no.get()
        add_name = result_name.get()
        add_desc = result_desc.get()
        
        if len(add_ptrno)==0 and len(add_name)==0 and len(add_desc) == 0:
            answer=messagebox.askokcancel("Question","Complete your data.")
            print(answer)
        else:
            
                
            # Open the workbook
            try:
                workbook = openpyxl.load_workbook('./resources/user/data/items.xlsx')
                worksheet = workbook.active
            except FileNotFoundError:
                workbook = Workbook()
                worksheet = workbook.active
            
                
            # Check the number of entries
            if len(worksheet['A']) - 1 >= 16:
                messagebox.showwarning("Limit Reached", "The maximum number of entries has been reached.")
                return

            # Check if the user already exists in the treeview
            for child in item_table_report.get_children():
                if add_ptrno == item_table_report.item(child)['values'][1]:
                    messagebox.showwarning("Item Exists", "The item already exists in the table.")
                    return
            # Find the next available row
            next_row = worksheet.max_row + 1
            #columns = ('date_acquired', 'property_no', 'name',  'description', 'quantity', 'condition')
            # Add the entry to the worksheet
            worksheet.cell(row=next_row, column=1).value = date.get()
            worksheet.cell(row=next_row, column=2).value = add_ptrno
            worksheet.cell(row=next_row, column=3).value = add_name
            worksheet.cell(row=next_row, column=4).value = add_desc
            worksheet.cell(row=next_row, column=5).value = quantity.get()
            worksheet.cell(row=next_row, column=6).value = condition.get()
            worksheet.cell(row=next_row, column=7).value = location_entry.get()
            
            # Save the workbook
            workbook.save('./resources/user/data/items.xlsx')

            # Clear the entry fields
            result_property_no.delete(0, len(add_ptrno))
            result_name.delete(0, len(add_name))
            result_desc.delete(0, len(add_desc))

            # Refresh the treeview
            refresh_treeview()
    
    def delete_entry():
        selected_item = item_table_report.selection()
        if selected_item:
            # Retrieve the entry ID from the selected item
            entry_id = item_table_report.item(selected_item)['values'][0]

            # Open the workbook
            workbook = openpyxl.load_workbook('./resources/user/data/items.xlsx')
            worksheet = workbook.active

            # Find the row with the matching entry ID
            for row in worksheet.iter_rows(min_row=2, values_only=True):
                if row[1] == entry_id:
                    # Get the row index from the treeview
                    index = int(item_table_report.index(selected_item)) + 1

                    # Delete the row from the worksheet
                    worksheet.delete_rows(index)
                    break
            
            # Save the workbook
            workbook.save('./resources/user/data/items.xlsx')

            # Refresh the treeview
            refresh_treeview()


    def delete_all_entries():
        # Open the workbook
        workbook = openpyxl.load_workbook('./resources/user/data/items.xlsx')
        worksheet = workbook.active

        # Clear all data rows except the header row
        worksheet.delete_rows(1, worksheet.max_row)

        # Save the workbook
        workbook.save('./resources/user/data/items.xlsx')

        # Refresh the treeview
        refresh_treeview()
    
    def refresh_treeview():
        # Clear existing treeview items
        for item in item_table_report.get_children():
            item_table_report.delete(item)

        # Load the workbook
        try:
            workbook = openpyxl.load_workbook('./resources/user/data/items.xlsx')
            worksheet = workbook.active

            # Retrieve data from the worksheet
            data = []
            for row in worksheet.iter_rows(min_row=2, values_only=True):
                data.append(row)

            # Insert data into the treeview
            for entry in data:
                item_table_report.insert("", len(entry), values=entry)
        except FileNotFoundError:
            pass
        
    
    
    def search_item():
        itemno = property_no.get()
        
        
        if len(itemno)== 0:
            answer=messagebox.askokcancel("Question","Enter Property No. first.")
            print(answer)
        else:    
            
            try:
                item_db = db.child("items_info").order_by_child("NO").equal_to(itemno).get()
                
                for dt in item_db.each():
                    var1 = dt.val()['NO']
                    var2 = dt.val()['NAME']
                    var3 = dt.val()['DES']
                    var4 = dt.val()['LOC']
                    var5 = dt.val()['QTY']
                    var6 = dt.val()['CON']
                
                result_property_no.delete(0,len(var1))
                result_property_no.insert(0,var1)
                result_name.delete(0,len(var2))
                result_name.insert(0,var2)
                result_desc.delete(0,len(var3))
                result_desc.insert(0,var3)
                location_entry.delete(0,len(var4))
                location_entry.insert(0,var4)
                qty_var.set(var5)
                condition.set(var6)
                
                
            except:
                answer=messagebox.askokcancel("Question","This item is not exist.")
                print(answer)    
    
        
    def generate_report():
        
        ptrno = ptr_no.get()
        fc = fund_cluster.get()
        dt = date.get()
        fromao = from_ao_afc.get()
        toao = to_ao_afc.get()
        one = transfer_var1.get()
        two = transfer_var2.get()
        three = transfer_var3.get()
        four = transfer_var4.get()
        other = transfer_4_entry.get()
        
        if len(ptrno) == 0:
            answer=messagebox.askokcancel("HELLO","Please add the PTR No before generating.")
            print(answer) 
        else:

                wb = load_workbook("./resources/user/data/UserLog.xlsx", data_only=True)
                sh = wb["Report"]
                L1 = sh["A1"].value
                L2 = sh["A2"].value
                L3 = sh["A3"].value
                
                if L1 == None:
                    L1 = ''
                if L2 == None:
                    L2 = ''
                if L3== None:
                    L3 = ''
                
                
                wb2 = load_workbook("./resources/user/data/items.xlsx", data_only=True)
                sh2 = wb2["Sheet"]
                A2 = sh2["A2"].value
                B2 = sh2["B2"].value
                C2 = sh2["C2"].value
                D2 = sh2["D2"].value
                E2 = sh2["E2"].value
                F2 = sh2["F2"].value
                G2 = sh2["G2"].value
                A3 = sh2["A3"].value
                B3 = sh2["B3"].value
                C3 = sh2["C3"].value
                D3 = sh2["D3"].value
                E3 = sh2["E3"].value
                F3 = sh2["F3"].value
                G3 = sh2["G3"].value
                A4 = sh2["A4"].value
                B4 = sh2["B4"].value
                C4 = sh2["C4"].value
                D4 = sh2["D4"].value
                E4 = sh2["E4"].value
                F4 = sh2["F4"].value
                G4 = sh2["G4"].value
                A5 = sh2["A5"].value
                B5 = sh2["B5"].value
                C5 = sh2["C5"].value
                D5 = sh2["D5"].value
                E5 = sh2["E5"].value
                F5 = sh2["F5"].value
                G5 = sh2["G5"].value
                A6 = sh2["A6"].value
                B6 = sh2["B6"].value
                C6 = sh2["C6"].value
                D6 = sh2["D6"].value
                E6 = sh2["E6"].value
                F6 = sh2["F6"].value
                G6 = sh2["G6"].value
                A7 = sh2["A7"].value
                B7 = sh2["B7"].value
                C7 = sh2["C7"].value
                D7 = sh2["D7"].value
                E7 = sh2["E7"].value
                F7 = sh2["F7"].value
                G7 = sh2["G7"].value
                A8 = sh2["A8"].value
                B8 = sh2["B8"].value
                C8 = sh2["C8"].value
                D8 = sh2["D8"].value
                E8 = sh2["E8"].value
                F8 = sh2["F8"].value
                G8 = sh2["G8"].value
                A9 = sh2["A9"].value
                B9 = sh2["B9"].value
                C9 = sh2["C9"].value
                D9 = sh2["D9"].value
                E9 = sh2["E9"].value
                F9 = sh2["F9"].value
                G9 = sh2["G9"].value
                A10 = sh2["A10"].value
                B10 = sh2["B10"].value
                C10 = sh2["C10"].value
                D10 = sh2["D10"].value
                E10 = sh2["E10"].value
                F10 = sh2["F10"].value
                G10 = sh2["G10"].value
                A11 = sh2["A11"].value
                B11 = sh2["B11"].value
                C11 = sh2["C11"].value
                D11 = sh2["D11"].value
                E11 = sh2["E11"].value
                F11 = sh2["F11"].value
                G11 = sh2["G11"].value
                A12 = sh2["A12"].value
                B12 = sh2["B12"].value
                C12 = sh2["C12"].value
                D12 = sh2["D12"].value
                E12 = sh2["E12"].value
                F12 = sh2["F12"].value
                G12 = sh2["G12"].value
                A13 = sh2["A13"].value
                B13 = sh2["B13"].value
                C13 = sh2["C13"].value
                D13 = sh2["D13"].value
                E13 = sh2["E13"].value
                F13 = sh2["F13"].value
                G13 = sh2["G13"].value
                A14 = sh2["A14"].value
                B14 = sh2["B14"].value
                C14 = sh2["C14"].value
                D14 = sh2["D14"].value
                E14 = sh2["E14"].value
                F14 = sh2["F14"].value
                G14 = sh2["G14"].value
                A15 = sh2["A15"].value
                B15 = sh2["B15"].value
                C15 = sh2["C15"].value
                D15 = sh2["D15"].value
                E15 = sh2["E15"].value
                F15 = sh2["F15"].value
                G15 = sh2["G15"].value
                A16 = sh2["A16"].value
                B16 = sh2["B16"].value
                C16 = sh2["C16"].value
                D16 = sh2["D16"].value
                E16 = sh2["E16"].value
                F16 = sh2["F16"].value
                G16 = sh2["G16"].value
                A17 = sh2["A17"].value
                B17 = sh2["B17"].value
                C17 = sh2["C17"].value
                D17 = sh2["D17"].value
                E17 = sh2["E17"].value
                F17 = sh2["F17"].value
                G17 = sh2["G17"].value
                A18 = sh2["A18"].value
                B18 = sh2["B18"].value
                C18 = sh2["C18"].value
                D18 = sh2["D18"].value
                E18 = sh2["E18"].value
                F18 = sh2["F18"].value
                G18 = sh2["G18"].value
                
                #2
                if A2 == None:
                    A2 = ''
                else:
                    db.child("items_info").child(B2).update({"DTA":A2})
                    
                if B2 == None:
                    B2 = ''
                else:
                    db.child("items_info").child(B2).update({"NO":B2})
                    
                if C2== None:
                    C2 = ''
                else:
                    db.child("items_info").child(B2).update({"NAME":C2})
                    
                if D2== None:
                    D2 = ''
                else:
                    db.child("items_info").child(B2).update({"DES":D2})
                    
                if E2 == None:
                    E2 = ''
                else:
                    db.child("items_info").child(B2).update({"QTY":E2})
                    
                if F2 == None:
                    F2 = '' 
                else:
                    db.child("items_info").child(B2).update({"CON":F2})
                    
                if G2 == None:
                    G2 = '' 
                else:
                    db.child("items_info").child(B2).update({"LOC":G2})
                  
                #3    
                if A3 == None:
                    A3 = ''
                else:
                    db.child("items_info").child(B3).update({"DTA":A3})
                
                if B3 == None:
                    B3 = ''
                else:
                    db.child("items_info").child(B3).update({"NO":B3})
                
                if C3== None:
                    C3 = ''
                else:
                    db.child("items_info").child(B3).update({"NAME":C3})
                    
                if D3== None:
                    D3 = ''
                else:
                    db.child("items_info").child(B3).update({"DES":D3})
                    
                if E3 == None:
                    E3 = ''
                else:
                    db.child("items_info").child(B3).update({"QTY":E3})
                    
                if F3 == None:
                    F3 = ''  
                else:
                    db.child("items_info").child(B3).update({"CON":F3})
                    
                if G3 == None:
                    G3 = '' 
                else:
                    db.child("items_info").child(B3).update({"LOC":G3})
                    
                #4       
                if A4 == None:
                    A4 = ''
                else:
                    db.child("items_info").child(B4).update({"DTA":A4})
                
                if B4 == None:
                    B4 = ''
                else:
                    db.child("items_info").child(B4).update({"NO":B4})
                
                if C4== None:
                    C4 = ''
                else:
                    db.child("items_info").child(B4).update({"NAME":C4})
                    
                if D4== None:
                    D4 = ''
                else:
                    db.child("items_info").child(B4).update({"DES":D4})
                    
                if E4 == None:
                    E4 = ''
                else:
                    db.child("items_info").child(B4).update({"QTY":E4})
                    
                if F4 == None:
                    F4 = ''  
                else:
                    db.child("items_info").child(B4).update({"CON":F4})
                    
                if G4 == None:
                    G4 = '' 
                else:
                    db.child("items_info").child(B4).update({"LOC":G4})
                    
                #5    
                if A5 == None:
                    A5 = ''
                else:
                    db.child("items_info").child(B5).update({"DTA":A5})
                
                if B5 == None:
                    B5 = ''
                else:
                    db.child("items_info").child(B5).update({"NO":B5})
                
                if C5== None:
                    C5 = ''
                else:
                    db.child("items_info").child(B5).update({"NAME":C5})
                    
                if D5== None:
                    D5 = ''
                else:
                    db.child("items_info").child(B5).update({"DES":D5})
                    
                if E5 == None:
                    E5 = ''
                else:
                    db.child("items_info").child(B5).update({"QTY":E5})
                    
                if F5 == None:
                    F5 = ''  
                else:
                    db.child("items_info").child(B5).update({"CON":F4})
                    
                if G5 == None:
                    G5 = '' 
                else:
                    db.child("items_info").child(B5).update({"LOC":G5})
                
                #6    
                if A6 == None:
                    A6 = ''
                else:
                    db.child("items_info").child(B6).update({"DTA":A6})
                
                if B6 == None:
                    B6 = ''
                else:
                    db.child("items_info").child(B6).update({"NO":B6})
                
                if C6== None:
                    C6 = ''
                else:
                    db.child("items_info").child(B6).update({"NAME":C6})
                    
                if D6== None:
                    D6 = ''
                else:
                    db.child("items_info").child(B6).update({"DES":D6})
                    
                if E6 == None:
                    E6 = ''
                else:
                    db.child("items_info").child(B6).update({"QTY":E6})
                    
                if F6 == None:
                    F6 = ''  
                else:
                    db.child("items_info").child(B6).update({"CON":F6})
                    
                if G6 == None:
                    G6 = '' 
                else:
                    db.child("items_info").child(B6).update({"LOC":G6})
                #7    
                
                if A7 == None:
                    A7 = ''
                else:
                    db.child("items_info").child(B7).update({"DTA":A7})
                
                if B7 == None:
                    B7 = ''
                else:
                    db.child("items_info").child(B7).update({"NO":B7})
                
                if C7== None:
                    C7 = ''
                else:
                    db.child("items_info").child(B7).update({"NAME":C7})
                    
                if D7== None:
                    D7 = ''
                else:
                    db.child("items_info").child(B7).update({"DES":D7})
                    
                if E7 == None:
                    E7 = ''
                else:
                    db.child("items_info").child(B7).update({"QTY":E7})
                    
                if F7 == None:
                    F7 = ''  
                else:
                    db.child("items_info").child(B7).update({"CON":F7})
                    
                if G7 == None:
                    G7 = '' 
                else:
                    db.child("items_info").child(B7).update({"LOC":G7}) 
                #8    
                if A8 == None:
                    A8 = ''
                else:
                    db.child("items_info").child(B8).update({"DTA":A8})
                
                if B8 == None:
                    B8 = ''
                else:
                    db.child("items_info").child(B8).update({"NO":B8})
                
                if C8== None:
                    C8 = ''
                else:
                    db.child("items_info").child(B8).update({"NAME":C8})
                    
                if D8== None:
                    D8 = ''
                else:
                    db.child("items_info").child(B8).update({"DES":D8})
                  
                if E8 == None:
                    E8 = ''
                else:
                    db.child("items_info").child(B8).update({"QTY":E8})
                    
                if F8 == None:
                    F8 = ''  
                else:
                    db.child("items_info").child(B8).update({"CON":F8})
                    
                if G8 == None:
                    G8 = '' 
                else:
                    db.child("items_info").child(B8).update({"LOC":G8})   
                #9    
                if A9 == None:
                    A9 = ''
                else:
                    db.child("items_info").child(B9).update({"DTA":A9})
                
                if B9 == None:
                    B9 = ''
                else:
                    db.child("items_info").child(B9).update({"NO":B9})
                
                if C9== None:
                    C9 = ''
                else:
                    db.child("items_info").child(B9).update({"NAME":C9})
                    
                if D9== None:
                    D9 = ''
                else:
                    db.child("items_info").child(B9).update({"DES":D9})
                  
                if E9 == None:
                    E9 = ''
                else:
                    db.child("items_info").child(B9).update({"QTY":E9})
                    
                if F9 == None:
                    F9 = ''  
                else:
                    db.child("items_info").child(B9).update({"CON":F9})
                    
                if G9 == None:
                    G9 = '' 
                else:
                    db.child("items_info").child(B9).update({"LOC":G9}) 
                #10    
                if A10 == None:
                    A10 = ''
                else:
                    db.child("items_info").child(B10).update({"DTA":A10})
                
                if B10 == None:
                    B10 = ''
                else:
                    db.child("items_info").child(B10).update({"NO":B10})
                
                if C10== None:
                    C10 = ''
                else:
                    db.child("items_info").child(B10).update({"NAME":C10})
                    
                if D10== None:
                    D10 = ''
                else:
                    db.child("items_info").child(B10).update({"DES":D10})
                  
                if E10 == None:
                    E10 = ''
                else:
                    db.child("items_info").child(B10).update({"QTY":E10})
                    
                if F10 == None:
                    F10 = ''  
                else:
                    db.child("items_info").child(B10).update({"CON":F10})
                    
                if G10 == None:
                    G10 = '' 
                else:
                    db.child("items_info").child(B10).update({"LOC":G10})   
                #11
                if A11 == None:
                    A11 = ''
                else:
                    db.child("items_info").child(B11).update({"DTA":A11})
                
                if B11 == None:
                    B11 = ''
                else:
                    db.child("items_info").child(B11).update({"NO":B11})
                
                if C11== None:
                    C11 = ''
                else:
                    db.child("items_info").child(B11).update({"NAME":C11})
                    
                if D11== None:
                    D11 = ''
                else:
                    db.child("items_info").child(B11).update({"DES":D11})
                  
                if E11 == None:
                    E11 = ''
                else:
                    db.child("items_info").child(B11).update({"QTY":E11})
                    
                if F11 == None:
                    F11 = ''  
                else:
                    db.child("items_info").child(B11).update({"CON":F11})
                    
                if G11 == None:
                    G11 = '' 
                else:
                    db.child("items_info").child(B11).update({"LOC":G11})
                 #12
                if A12 == None:
                    A12 = ''
                else:
                    db.child("items_info").child(B12).update({"DTA":A12})
                    
                if B12 == None:
                    B12 = ''
                else:
                    db.child("items_info").child(B12).update({"NO":B12})
                    
                if C12 == None:
                    C12 = ''
                else:
                    db.child("items_info").child(B12).update({"NAME":C12})
                    
                if D12== None:
                    D12 = ''
                else:
                    db.child("items_info").child(B12).update({"DES":D12})
                    
                if E12 == None:
                    E12 = ''
                else:
                    db.child("items_info").child(B12).update({"QTY":E12})
                    
                if F12 == None:
                    F12 = '' 
                else:
                    db.child("items_info").child(B12).update({"CON":F12})
                    
                if G12 == None:
                    G12 = '' 
                else:
                    db.child("items_info").child(B12).update({"LOC":G12})
                  
                #13    
                if A13 == None:
                    A13 = ''
                else:
                    db.child("items_info").child(B13).update({"DTA":A13})
                
                if B13 == None:
                    B13 = ''
                else:
                    db.child("items_info").child(B13).update({"NO":B13})
                
                if C13== None:
                    C13 = ''
                else:
                    db.child("items_info").child(B13).update({"NAME":C13})
                    
                if D13== None:
                    D13 = ''
                else:
                    db.child("items_info").child(B13).update({"DES":D13})
                    
                if E13 == None:
                    E13 = ''
                else:
                    db.child("items_info").child(B13).update({"QTY":E13})
                    
                if F13 == None:
                    F13 = ''  
                else:
                    db.child("items_info").child(B13).update({"CON":F13})
                    
                if G13 == None:
                    G13 = '' 
                else:
                    db.child("items_info").child(B13).update({"LOC":G13})
                    
                #14       
                if A14 == None:
                    A14 = ''
                else:
                    db.child("items_info").child(B14).update({"DTA":A14})
                
                if B14 == None:
                    B14 = ''
                else:
                    db.child("items_info").child(B14).update({"NO":B14})
                
                if C14== None:
                    C14 = ''
                else:
                    db.child("items_info").child(B14).update({"NAME":C14})
                    
                if D14== None:
                    D14 = ''
                else:
                    db.child("items_info").child(B14).update({"DES":D14})
                    
                if E14 == None:
                    E14 = ''
                else:
                    db.child("items_info").child(B14).update({"QTY":E14})
                    
                if F14 == None:
                    F14 = ''  
                else:
                    db.child("items_info").child(B14).update({"CON":F14})
                    
                if G14 == None:
                    G14 = '' 
                else:
                    db.child("items_info").child(B14).update({"LOC":G14})
                    
                #15    
                if A15 == None:
                    A15 = ''
                else:
                    db.child("items_info").child(B15).update({"DTA":A15})
                
                if B15 == None:
                    B15 = ''
                else:
                    db.child("items_info").child(B15).update({"NO":B15})
                
                if C15 == None:
                    C15 = ''
                else:
                    db.child("items_info").child(B15).update({"NAME":C15})
                    
                if D15== None:
                    D15 = ''
                else:
                    db.child("items_info").child(B15).update({"DES":D15})
                    
                if E15 == None:
                    E15 = ''
                else:
                    db.child("items_info").child(B15).update({"QTY":E15})
                    
                if F15 == None:
                    F15 = ''  
                else:
                    db.child("items_info").child(B15).update({"CON":F15})
                    
                if G15 == None:
                    G15 = '' 
                else:
                    db.child("items_info").child(B15).update({"LOC":G15})
                
                #16    
                if A16 == None:
                    A16 = ''
                else:
                    db.child("items_info").child(B16).update({"DTA":A16})
                
                if B16 == None:
                    B16 = ''
                else:
                    db.child("items_info").child(B16).update({"NO":B16})
                
                if C16== None:
                    C16 = ''
                else:
                    db.child("items_info").child(B16).update({"NAME":C16})
                    
                if D16== None:
                    D16 = ''
                else:
                    db.child("items_info").child(B16).update({"DES":D16})
                if E16 == None:
                    E16 = ''
                else:
                    db.child("items_info").child(B16).update({"QTY":E16})
                    
                if F16 == None:
                    F16 = ''  
                else:
                    db.child("items_info").child(B16).update({"CON":F16})
                    
                if G16 == None:
                    G16 = '' 
                else:
                    db.child("items_info").child(B16).update({"LOC":G16})
                #17    
                
                if A17 == None:
                    A17 = ''
                else:
                    db.child("items_info").child(B17).update({"DTA":A17})
                
                if B17 == None:
                    B17 = ''
                else:
                    db.child("items_info").child(B17).update({"NO":B17})
                
                if C17== None:
                    C17 = ''
                else:
                    db.child("items_info").child(B17).update({"NAME":C17})
                    
                if D17== None:
                    D17 = ''
                else:
                    db.child("items_info").child(B17).update({"DES":D17})
                if E17 == None:
                    E17 = ''
                else:
                    db.child("items_info").child(B17).update({"QTY":E17})
                    
                if F17 == None:
                    F17 = ''  
                else:
                    db.child("items_info").child(B17).update({"CON":F17})
                    
                if G17 == None:
                    G17 = '' 
                else:
                    db.child("items_info").child(B17).update({"LOC":G17}) 
                #8    
                if A18 == None:
                    A18 = ''
                else:
                    db.child("items_info").child(B18).update({"DTA":A18})
                
                if B18 == None:
                    B18 = ''
                else:
                    db.child("items_info").child(B18).update({"NO":B18})
                
                if C18== None:
                    C18 = ''
                else:
                    db.child("items_info").child(B18).update({"NAME":C18})
                    
                if D18== None:
                    D18 = ''
                else:
                    db.child("items_info").child(B18).update({"DES":D18})
                if E18 == None:
                    E18 = ''
                else:
                    db.child("items_info").child(B18).update({"QTY":E18})
                    
                if F18 == None:
                    F18 = ''  
                else:
                    db.child("items_info").child(B18).update({"CON":F18})
                    
                if G18 == None:
                    G18 = '' 
                else:
                    db.child("items_info").child(B18).update({"LOC":G18}) 
                                        
                
                if one == "on":
                    
                    send_prt_info = {
                        'PTR': ptrno,
                        'FROM': fromao,
                        'TO': toao,
                        'FC': fc,
                        'DT': dt,
                        'TRANSFER':'Donation',
                        'REASON1':L1,
                        'REASON2': L2,
                        'REASON3': L3,
                        'DA1': A2,
                        'PN1': B2,
                        'DESC1': D2,
                        'AM1': E2,
                        'CON1': F2,
                        'LOC1': G2,
                        'DA2': A3,
                        'PN2': B3,
                        'DESC2': D3,
                        'AM2': E3,
                        'CON2': F3,
                        'LOC2': G3,
                        'DA3': A4,
                        'PN3': B4,
                        'DESC3': D4,
                        'AM3': E4,
                        'CON3': F4,
                        'LOC3': G4,
                        'DA4': A5,
                        'PN4': B5,
                        'DESC4': D5,
                        'AM4': E5,
                        'CON4': F5,
                        'LOC4': G5,
                        'DA5': A6,
                        'PN5': B6,
                        'DESC5': D6,
                        'AM5': E6,
                        'CON5': F6,
                        'LOC5': G6,
                        'DA6': A7,
                        'PN6': B7,
                        'DESC6': D7,
                        'AM6': E7,
                        'CON6': F7,
                        'LOC6': G7,
                        'DA7': A8,
                        'PN7': B8,
                        'DESC7': D8,
                        'AM7': E8,
                        'CON7': F8,
                        'LOC7': G8,
                        'DA8': A9,
                        'PN8': B9,
                        'DESC8': D9,
                        'AM8': E9,
                        'CON8': F9,
                        'LOC8': G9,
                        'DA9': A10,
                        'PN9': B10,
                        'DESC9': D10,
                        'AM9': E10,
                        'CON9': F10,
                        'LOC9': G10,
                        'DA10': A11,
                        'PN10': B11,
                        'DESC10': D11,
                        'AM10': E11,
                        'CON10': F11,
                        'LOC10': G11,
                        'DA11': A12,
                        'PN11': B12,
                        'DESC11': D12,
                        'AM11': E12,
                        'CON11': F12,
                        'LOC11': G12,
                        'DA12': A13,
                        'PN12': B13,
                        'DESC12': D13,
                        'AM12': E13,
                        'LOC12': G13,
                        'CON12': F13,
                        'DA13': A14,
                        'PN13': B14,
                        'DESC13': D14,
                        'AM13': E14,
                        'CON13': F14,
                        'LOC13': G14,
                        'DA14': A15,
                        'PN14': B15,
                        'DESC14': D15,
                        'AM14': E15,
                        'CON14': F15,
                        'LOC14': G15,
                        'DA15': A16,
                        'PN15': B16,
                        'DESC15': D16,
                        'AM15': E16,
                        'CON15': F16,
                        'LOC15': G16,
                        'DA16': A17,
                        'PN16': B17,
                        'DESC16': D17,
                        'AM16': E17,
                        'CON16': F17,
                        'LOC16': G17,
                        'DA17': A18,
                        'PN17': B18,
                        'DESC17': D18,
                        'AM17': E18,
                        'CON17': F18,
                        'LOC17': G18,
                        'NAME': approve_name.get(),
                        'RINAME':release_name.get(),
                        'RNAME': receive_name.get(),
                        'DES': approve_designation.get(),
                        'RIDES':release_designation.get(),
                        'RDES':receive_designation.get(),
                        'ADT': approve_date.get(),
                        'RIDT': release_date.get(),
                        'RDT':receive_date.get(),
                        'CREATED': current_day+" ("+current_time+")",
                        'EMAIL': email,
                        'REMARKS': 'PENDING',
                        'N1': C2,
                        'N2': C3,
                        'N3': C4,
                        'N4': C5,
                        'N5': C6,
                        'N6': C7,
                        'N7': C8,
                        'N8': C9,
                        'N9': C10,
                        'N10': C11,
                        'N11': C12,
                        'N12': C13,
                        'N13': C14,
                        'N14': C15,
                        'N15': C16,
                        'N16': C17,
                        'N17': C18
                        
                    }
                    
                    db.child("generated_files").child(ptrno).set(send_prt_info)
                    
                    answer=messagebox.askokcancel("HELLO","Your PTR is submitted successfully , Please wait for approval.")
                    print(answer)
                    
                if two == "on":
                    send_prt_info = {
                        'PTR': ptrno,
                        'FROM': fromao,
                        'TO': toao,
                        'FC': fc,
                        'DT': dt,
                        'TRANSFER':'Relocate',
                        'REASON1':L1,
                        'REASON2': L2,
                        'REASON3': L3,
                        'DA1': A2,
                        'PN1': B2,
                        'DESC1': D2,
                        'AM1': E2,
                        'CON1': F2,
                        'LOC1': G2,
                        'DA2': A3,
                        'PN2': B3,
                        'DESC2': D3,
                        'AM2': E3,
                        'CON2': F3,
                        'LOC2': G3,
                        'DA3': A4,
                        'PN3': B4,
                        'DESC3': D4,
                        'AM3': E4,
                        'CON3': F4,
                        'LOC3': G4,
                        'DA4': A5,
                        'PN4': B5,
                        'DESC4': D5,
                        'AM4': E5,
                        'CON4': F5,
                        'LOC4': G5,
                        'DA5': A6,
                        'PN5': B6,
                        'DESC5': D6,
                        'AM5': E6,
                        'CON5': F6,
                        'LOC5': G6,
                        'DA6': A7,
                        'PN6': B7,
                        'DESC6': D7,
                        'AM6': E7,
                        'CON6': F7,
                        'LOC6': G7,
                        'DA7': A8,
                        'PN7': B8,
                        'DESC7': D8,
                        'AM7': E8,
                        'CON7': F8,
                        'LOC7': G8,
                        'DA8': A9,
                        'PN8': B9,
                        'DESC8': D9,
                        'AM8': E9,
                        'CON8': F9,
                        'LOC8': G9,
                        'DA9': A10,
                        'PN9': B10,
                        'DESC9': D10,
                        'AM9': E10,
                        'CON9': F10,
                        'LOC9': G10,
                        'DA10': A11,
                        'PN10': B11,
                        'DESC10': D11,
                        'AM10': E11,
                        'CON10': F11,
                        'LOC10': G11,
                        'DA11': A12,
                        'PN11': B12,
                        'DESC11': D12,
                        'AM11': E12,
                        'CON11': F12,
                        'LOC11': G12,
                        'DA12': A13,
                        'PN12': B13,
                        'DESC12': D13,
                        'AM12': E13,
                        'LOC12': G13,
                        'CON12': F13,
                        'DA13': A14,
                        'PN13': B14,
                        'DESC13': D14,
                        'AM13': E14,
                        'CON13': F14,
                        'LOC13': G14,
                        'DA14': A15,
                        'PN14': B15,
                        'DESC14': D15,
                        'AM14': E15,
                        'CON14': F15,
                        'LOC14': G15,
                        'DA15': A16,
                        'PN15': B16,
                        'DESC15': D16,
                        'AM15': E16,
                        'CON15': F16,
                        'LOC15': G16,
                        'DA16': A17,
                        'PN16': B17,
                        'DESC16': D17,
                        'AM16': E17,
                        'CON16': F17,
                        'LOC16': G17,
                        'DA17': A18,
                        'PN17': B18,
                        'DESC17': D18,
                        'AM17': E18,
                        'CON17': F18,
                        'LOC17': G18,
                        'NAME': approve_name.get(),
                        'RINAME':release_name.get(),
                        'RNAME': receive_name.get(),
                        'DES': approve_designation.get(),
                        'RIDES':release_designation.get(),
                        'RDES':receive_designation.get(),
                        'ADT': approve_date.get(),
                        'RIDT': release_date.get(),
                        'RDT':receive_date.get(),
                        'CREATED': current_day+" ("+current_time+")",
                        'EMAIL': email,
                        'REMARKS': 'PENDING',
                        'N1': C2,
                        'N2': C3,
                        'N3': C4,
                        'N4': C5,
                        'N5': C6,
                        'N6': C7,
                        'N7': C8,
                        'N8': C9,
                        'N9': C10,
                        'N10': C11,
                        'N11': C12,
                        'N12': C13,
                        'N13': C14,
                        'N14': C15,
                        'N15': C16,
                        'N16': C17,
                        'N17': C18
                        
                    }
                    db.child("generated_files").child(ptrno).set(send_prt_info)
                    answer=messagebox.askokcancel("HELLO","Your PTR is submitted successfully , Please wait for approval.")
                    
                if three == "on":
                    send_prt_info = {
                        'PTR': ptrno,
                        'FROM': fromao,
                        'TO': toao,
                        'FC': fc,
                        'DT': dt,
                        'TRANSFER':'Reassignment ',
                        'REASON1':L1,
                        'REASON2': L2,
                        'REASON3': L3,
                        'DA1': A2,
                        'PN1': B2,
                        'DESC1': D2,
                        'AM1': E2,
                        'CON1': F2,
                        'LOC1': G2,
                        'DA2': A3,
                        'PN2': B3,
                        'DESC2': D3,
                        'AM2': E3,
                        'CON2': F3,
                        'LOC2': G3,
                        'DA3': A4,
                        'PN3': B4,
                        'DESC3': D4,
                        'AM3': E4,
                        'CON3': F4,
                        'LOC3': G4,
                        'DA4': A5,
                        'PN4': B5,
                        'DESC4': D5,
                        'AM4': E5,
                        'CON4': F5,
                        'LOC4': G5,
                        'DA5': A6,
                        'PN5': B6,
                        'DESC5': D6,
                        'AM5': E6,
                        'CON5': F6,
                        'LOC5': G6,
                        'DA6': A7,
                        'PN6': B7,
                        'DESC6': D7,
                        'AM6': E7,
                        'CON6': F7,
                        'LOC6': G7,
                        'DA7': A8,
                        'PN7': B8,
                        'DESC7': D8,
                        'AM7': E8,
                        'CON7': F8,
                        'LOC7': G8,
                        'DA8': A9,
                        'PN8': B9,
                        'DESC8': D9,
                        'AM8': E9,
                        'CON8': F9,
                        'LOC8': G9,
                        'DA9': A10,
                        'PN9': B10,
                        'DESC9': D10,
                        'AM9': E10,
                        'CON9': F10,
                        'LOC9': G10,
                        'DA10': A11,
                        'PN10': B11,
                        'DESC10': D11,
                        'AM10': E11,
                        'CON10': F11,
                        'LOC10': G11,
                        'DA11': A12,
                        'PN11': B12,
                        'DESC11': D12,
                        'AM11': E12,
                        'CON11': F12,
                        'LOC11': G12,
                        'DA12': A13,
                        'PN12': B13,
                        'DESC12': D13,
                        'AM12': E13,
                        'LOC12': G13,
                        'CON12': F13,
                        'DA13': A14,
                        'PN13': B14,
                        'DESC13': D14,
                        'AM13': E14,
                        'CON13': F14,
                        'LOC13': G14,
                        'DA14': A15,
                        'PN14': B15,
                        'DESC14': D15,
                        'AM14': E15,
                        'CON14': F15,
                        'LOC14': G15,
                        'DA15': A16,
                        'PN15': B16,
                        'DESC15': D16,
                        'AM15': E16,
                        'CON15': F16,
                        'LOC15': G16,
                        'DA16': A17,
                        'PN16': B17,
                        'DESC16': D17,
                        'AM16': E17,
                        'CON16': F17,
                        'LOC16': G17,
                        'DA17': A18,
                        'PN17': B18,
                        'DESC17': D18,
                        'AM17': E18,
                        'CON17': F18,
                        'LOC17': G18,
                        'NAME': approve_name.get(),
                        'RINAME':release_name.get(),
                        'RNAME': receive_name.get(),
                        'DES': approve_designation.get(),
                        'RIDES':release_designation.get(),
                        'RDES':receive_designation.get(),
                        'ADT': approve_date.get(),
                        'RIDT': release_date.get(),
                        'RDT':receive_date.get(),
                        'CREATED': current_day+" ("+current_time+")",
                        'EMAIL': email,
                        'REMARKS': 'PENDING',
                        'N1': C2,
                        'N2': C3,
                        'N3': C4,
                        'N4': C5,
                        'N5': C6,
                        'N6': C7,
                        'N7': C8,
                        'N8': C9,
                        'N9': C10,
                        'N10': C11,
                        'N11': C12,
                        'N12': C13,
                        'N13': C14,
                        'N14': C15,
                        'N15': C16,
                        'N16': C17,
                        'N17': C18
                        
                    }
                    db.child("generated_files").child(ptrno).set(send_prt_info)
                    answer=messagebox.askokcancel("HELLO","Your PTR is submitted successfully , Please wait for approval.")
                if four == "on":
                    send_prt_info = {
                        'PTR': ptrno,
                        'FROM': fromao,
                        'TO': toao,
                        'FC': fc,
                        'DT': dt,
                        'TRANSFER':transfer_4_entry.get().title(),
                        'REASON1':L1,
                        'REASON2': L2,
                        'REASON3': L3,
                        'DA1': A2,
                        'PN1': B2,
                        'DESC1': D2,
                        'AM1': E2,
                        'CON1': F2,
                        'LOC1': G2,
                        'DA2': A3,
                        'PN2': B3,
                        'DESC2': D3,
                        'AM2': E3,
                        'CON2': F3,
                        'LOC2': G3,
                        'DA3': A4,
                        'PN3': B4,
                        'DESC3': D4,
                        'AM3': E4,
                        'CON3': F4,
                        'LOC3': G4,
                        'DA4': A5,
                        'PN4': B5,
                        'DESC4': D5,
                        'AM4': E5,
                        'CON4': F5,
                        'LOC4': G5,
                        'DA5': A6,
                        'PN5': B6,
                        'DESC5': D6,
                        'AM5': E6,
                        'CON5': F6,
                        'LOC5': G6,
                        'DA6': A7,
                        'PN6': B7,
                        'DESC6': D7,
                        'AM6': E7,
                        'CON6': F7,
                        'LOC6': G7,
                        'DA7': A8,
                        'PN7': B8,
                        'DESC7': D8,
                        'AM7': E8,
                        'CON7': F8,
                        'LOC7': G8,
                        'DA8': A9,
                        'PN8': B9,
                        'DESC8': D9,
                        'AM8': E9,
                        'CON8': F9,
                        'LOC8': G9,
                        'DA9': A10,
                        'PN9': B10,
                        'DESC9': D10,
                        'AM9': E10,
                        'CON9': F10,
                        'LOC9': G10,
                        'DA10': A11,
                        'PN10': B11,
                        'DESC10': D11,
                        'AM10': E11,
                        'CON10': F11,
                        'LOC10': G11,
                        'DA11': A12,
                        'PN11': B12,
                        'DESC11': D12,
                        'AM11': E12,
                        'CON11': F12,
                        'LOC11': G12,
                        'DA12': A13,
                        'PN12': B13,
                        'DESC12': D13,
                        'AM12': E13,
                        'LOC12': G13,
                        'CON12': F13,
                        'DA13': A14,
                        'PN13': B14,
                        'DESC13': D14,
                        'AM13': E14,
                        'CON13': F14,
                        'LOC13': G14,
                        'DA14': A15,
                        'PN14': B15,
                        'DESC14': D15,
                        'AM14': E15,
                        'CON14': F15,
                        'LOC14': G15,
                        'DA15': A16,
                        'PN15': B16,
                        'DESC15': D16,
                        'AM15': E16,
                        'CON15': F16,
                        'LOC15': G16,
                        'DA16': A17,
                        'PN16': B17,
                        'DESC16': D17,
                        'AM16': E17,
                        'CON16': F17,
                        'LOC16': G17,
                        'DA17': A18,
                        'PN17': B18,
                        'DESC17': D18,
                        'AM17': E18,
                        'CON17': F18,
                        'LOC17': G18,
                        'NAME': approve_name.get(),
                        'RINAME':release_name.get(),
                        'RNAME': receive_name.get(),
                        'DES': approve_designation.get(),
                        'RIDES':release_designation.get(),
                        'RDES':receive_designation.get(),
                        'ADT': approve_date.get(),
                        'RIDT': release_date.get(),
                        'RDT':receive_date.get(),
                        'CREATED': current_day+" ("+current_time+")",
                        'EMAIL': email,
                        'REMARKS': 'PENDING',
                        'N1': C2,
                        'N2': C3,
                        'N3': C4,
                        'N4': C5,
                        'N5': C6,
                        'N6': C7,
                        'N7': C8,
                        'N8': C9,
                        'N9': C10,
                        'N10': C11,
                        'N11': C12,
                        'N12': C13,
                        'N13': C14,
                        'N14': C15,
                        'N15': C16,
                        'N16': C17,
                        'N17': C18
                        
                    }
                    db.child("generated_files").child(ptrno).set(send_prt_info)
                    answer=messagebox.askokcancel("HELLO","Your PTR is submitted successfully , Please wait for approval.")
                    
                
            
       
        
        
        
        
            
    
    main = ctk.CTkFrame(report_form)
    main.pack(fill = 'both')
    #header frame
    header_frame = ctk.CTkFrame(main, fg_color='#313131', height=50, width=report_form.winfo_width(), corner_radius=0)
    header_frame.pack(fill = 'both', side='top')

    #app logo
    app_logo = ctk.CTkImage(Image.open('resources\images\DEFAULT\PTRLogo_White.png'), size=(25, 25))
    app_logo_label = ctk.CTkLabel(header_frame, image= app_logo, text=None)
    app_logo_label.pack(side = 'left', pady = 10, padx = 10)

    #app name
    app_name1 = ctk.CTkLabel(header_frame, text='PROPERTY', font= ctk.CTkFont('Arial', size = 22, weight = "bold"), text_color='white')
    app_name1.pack(side = 'left', pady = 10, padx = 10)
    app_name2 = ctk.CTkLabel(header_frame, text='TRANSFER', font= ctk.CTkFont('Arial', size = 22, weight = "bold"), text_color='#ee4444')
    app_name2.pack(side = 'left', pady = 10)

    label = ctk.CTkLabel(header_frame, text = 'Reports', font= ctk.CTkFont('Arial', size = 25, weight = "bold"), text_color='white')
    label.pack(side = 'right',  pady = 10, padx = 10)
        
    #notebook
    notebook = ttk.Notebook(main)
    notebook.pack(expand=True)
    
    # create frames
    tab1 = ctk.CTkFrame(notebook, width=1000, height=500, fg_color='white')
    tab2 = ctk.CTkFrame(notebook, width=1000, height=500, fg_color='white')
    tab3 = ctk.CTkFrame(notebook, width=1000, height=500, fg_color='white')

    tab1.pack(fill='both', expand=True)
    tab2.pack(fill='both', expand=True)
    tab3.pack(fill ='both', expand =True)

    # add frames to notebook
    notebook.add(tab1, text='PTR Information')
    notebook.add(tab2, text='Items')
    notebook.add(tab3, text='Approval')
    
    #first tab
    frame = ctk.CTkFrame(tab1, height=400, corner_radius=20, fg_color='#F4F4F4')
    frame.pack(fill = 'both', pady = 5, padx = 10)
    
    sub_frame = ctk.CTkFrame(frame, fg_color='transparent')
    sub_frame.pack(fill = 'x')
    
    entity_label = ctk.CTkLabel(sub_frame, text='Entity Name: ')
    entity_label.pack(side = 'left', pady = 5, padx = 10)
    
    entities = ["Polytechnic University of the Philippines - Lopez Branch"]
    entity_name = ctk.CTkOptionMenu(sub_frame, width=300, values=entities, fg_color= 'white', text_color='black', dropdown_fg_color='white', font= ctk.CTkFont('Arial', size = 12), dropdown_font=ctk.CTkFont('Arial', size = 12))
    entity_name.pack(side = 'left', pady = 5, padx = 10)
    
    fund_cluster_label = ctk.CTkLabel(sub_frame, text='Fund Cluster: ')
    fund_cluster_label.pack(side = 'left', pady = 5, padx = (170,0))
    fund_cluster= ctk.CTkEntry(sub_frame, fg_color='white')
    fund_cluster.pack(side = 'left', pady = 5, padx = 10)
    
    sub_frame2 = ctk.CTkFrame(frame, fg_color='transparent')
    sub_frame2.pack(fill = 'x')
    
    from_label = ctk.CTkLabel(sub_frame2, text='From Accountable Officer/Agency Cluster Fund: ')
    from_label.pack(side = 'left', pady = 5, padx = 10)
    from_ao_afc = ctk.CTkEntry(sub_frame2, width = 300, fg_color='white')
    from_ao_afc.pack(side = 'left', pady = 5, padx = 10)
    
    ptr_label = ctk.CTkLabel(sub_frame2, text='PTR No.: ')
    ptr_label.pack(side = 'left', pady = 5, padx = 10)
    ptr_no = ctk.CTkEntry(sub_frame2, fg_color='white', width = 155)
    ptr_no.pack(side = 'left', pady = 5, padx = 10)
    
    sub_frame3 = ctk.CTkFrame(frame, fg_color='transparent')
    sub_frame3.pack(fill = 'x')
    
    to_label = ctk.CTkLabel(sub_frame3, text='To Accountable Officer/Agency Cluster Fund: ')
    to_label.pack(side = 'left', pady = 5, padx = 10)
    to_ao_afc = ctk.CTkEntry(sub_frame3, width = 300, fg_color='white')
    to_ao_afc.pack(side = 'left', pady = 5, padx = 10)
    
    date_label = ctk.CTkLabel(sub_frame3, text='Date: ')
    date_label.pack(side = 'left', pady = 5, padx = 24)
    date = DateEntry(sub_frame3, font= ctk.CTkFont('Arial', size = 15) )
    date.pack(side = 'left', pady = 5, padx = 10, ipady =2, ipadx = 2)
    
    #transfer types
    frame2 = ctk.CTkFrame(tab1, corner_radius=20, fg_color='#F4F4F4')
    frame2.pack(fill = 'both', pady = 5, padx = 10)
    
    sub_frame4 = ctk.CTkFrame(frame2, fg_color='transparent')
    sub_frame4.pack(fill = 'x')
    
    transfer_type_label = ctk.CTkLabel(sub_frame4, text='Transfer Type (Select Only One): ')
    transfer_type_label.pack(side = 'left', pady = 10, padx = 10)
    
    
    transfer_var1 = ctk.StringVar()
    transfer_var2 = ctk.StringVar()
    transfer_var3 = ctk.StringVar()
    transfer_var4 = ctk.StringVar()
    transfer_var1.set("on")
    
    def checkbox_visibility1():
        if transfer_var1.get() == "on":
            transfer_var2.set("off")
            transfer_var3.set("off")
            transfer_var4.set("off")
    
    def checkbox_visibility2():
        if transfer_var2.get() == "on":
            transfer_var1.set("off")
            transfer_var3.set("off")
            transfer_var4.set("off")    
    
    def checkbox_visibility3():
        if transfer_var3.get() == "on":
            transfer_var1.set("off")
            transfer_var2.set("off")
            transfer_var4.set("off")
    
    def checkbox_visibility4():
        
        if transfer_var4.get() == "on":
            transfer_var1.set("off")
            transfer_var2.set("off")
            transfer_var3.set("off")
            transfer_4_entry.pack(side='left', pady=5, padx=10)
        else:
            transfer_4_entry.pack_forget()
 

    transfer_1 = ctk.CTkCheckBox(sub_frame4, command=checkbox_visibility1, text='Donation', variable=transfer_var1, onvalue="on", offvalue="off")
    transfer_1.pack(side='left', pady=5, padx=10)

    transfer_2 = ctk.CTkCheckBox(sub_frame4, command=checkbox_visibility2, text='Relocate', variable=transfer_var2, onvalue="on", offvalue="off")
    transfer_2.pack(side='left', pady=5, padx=10)

    transfer_3 = ctk.CTkCheckBox(sub_frame4, command=checkbox_visibility3, text='Reassign', variable=transfer_var3, onvalue="on", offvalue="off")
    transfer_3.pack(side='left', pady=5, padx=10)

    transfer_4 = ctk.CTkCheckBox(sub_frame4,command=checkbox_visibility4, text='Others (Specify): ',  variable=transfer_var4, onvalue="on", offvalue='off')
    transfer_4.pack(side='left', pady=5, padx=10)

    transfer_4_entry = ctk.CTkEntry(sub_frame4, fg_color='white')

    #second tab
    #items
    frame3 = ctk.CTkFrame(tab2,  fg_color='#F4F4F4')
    frame3.pack(fill = 'both', pady = 5, padx = 10)
    
    sub_frame4_3 = ctk.CTkFrame(frame3, fg_color='#F4F4F4')
    sub_frame4_3.pack(fill = 'x')
    
    
    
    item_label = ctk.CTkLabel(sub_frame4_3, text='Items')
    item_label.pack(side = 'left', pady = 5, padx = 10)

    
    
    #search item by property number
    selected_item = ctk.StringVar()
    search_icon = ctk.CTkImage(Image.open('resources\images\SETTING\search_icon.png'), size=(25, 25))
    search_button = ctk.CTkButton(sub_frame4_3,command= search_item, image=search_icon, text=None, width = 25, fg_color='transparent')
    search_button.pack(side = 'right', padx = (5, 10))
    sub_frame4_3_2 = ctk.CTkFrame(sub_frame4_3, fg_color='#F4F4F4')
    sub_frame4_3_2.pack(side = 'right')
    property_no = ctk.CTkEntry(sub_frame4_3_2, placeholder_text='Search Item Details', width = 200, fg_color='white')
    property_no.bind("<KeyRelease>", search_items)
    property_no.pack(side = 'right')
    
    #result of search
    sub_frame4_4 = ctk.CTkFrame(frame3, fg_color='#F4F4F4')
    sub_frame4_4.pack(fill = 'x', pady = 5)
    
    qty_var= ctk.StringVar()
    qty_var.set("1")
    result_label = ctk.CTkLabel(sub_frame4_4, text='Result: ')
    result_label.pack(side = 'left', pady = 5, padx = 10)
    result_property_no = ctk.CTkEntry(sub_frame4_4, placeholder_text='Property Number', width = 170, fg_color='white')
    result_property_no.pack(side = 'left', pady = 5, padx = 10)
    result_name = ctk.CTkEntry(sub_frame4_4, placeholder_text='Name', width = 170, fg_color='white')
    result_name.pack(side = 'left', pady = 5, padx = 10)
    result_desc = ctk.CTkEntry(sub_frame4_4, placeholder_text='Description', width = 170, fg_color='white')
    result_desc.pack(side = 'left', pady = 5, padx = 10)
    quantity = tk.Spinbox(sub_frame4_4, from_= 0, to= 1, borderwidth=1, font= ctk.CTkFont('Arial', size = 16), bg='white', textvariable=qty_var)
    quantity.pack(side = 'left', pady = 5, padx= 10)
    add_button = ctk.CTkButton(sub_frame4_4,command=add_entry, text='Add', width = 60, font=ctk.CTkFont('Arial', weight='bold'),fg_color='#0077B8', text_color='white')
    add_button.pack(side = 'left', pady = 5, padx = 10)
    
    #result of search additional
    sub_frame4_4_2 = ctk.CTkFrame(frame3, fg_color='#F4F4F4')
    sub_frame4_4_2.pack(fill = 'x', pady = 5)
    
    delete_button = ctk.CTkButton(sub_frame4_4_2,command=delete_entry, text='Delete', width = 60, font=ctk.CTkFont('Arial', weight='bold'),fg_color='#0077B8', text_color='white')
    delete_button.pack(side = 'right', pady = 5, padx = 10)
    
    deleteall_button = ctk.CTkButton(sub_frame4_4_2,command=delete_all_entries, text='Delete All', width = 60, font=ctk.CTkFont('Arial', weight='bold'),fg_color='#0077B8', text_color='white')
    deleteall_button.pack(side = 'right', pady = 5, padx = 10)
    
    condition_op = ["Serviceable", "Unserviceable"]
    condition = ctk.CTkOptionMenu(sub_frame4_4_2, width=250, values=condition_op, fg_color= 'white', text_color='black', dropdown_fg_color='white', font= ctk.CTkFont('Arial', size = 15), dropdown_font=ctk.CTkFont('Arial', size = 15))
    condition.pack(side = 'right', pady = 5, padx = 10)
    condition_label = ctk.CTkLabel(sub_frame4_4_2, text='Condition: ')
    condition_label.pack(side = 'right', pady = 5, padx = 10)
    
    selected_location = tk.StringVar()
    frame_loc = ctk.CTkFrame(sub_frame4_4_2, fg_color='#F4F4F4')
    frame_loc.pack(side = 'right')
    
    location_entry = ctk.CTkEntry(frame_loc, placeholder_text='Search Location', width = 200, fg_color='white')
    location_entry.bind("<KeyRelease>", search_locations)
    location_entry.pack(side = 'right', pady = 5, padx = 10)
    
    #item table
    sub_frame4_5 = ctk.CTkFrame(tab2, fg_color='#F4F4F4')
    sub_frame4_5.pack(fill = 'x', pady = 5, padx = 10)
    
    columns = ('date_acquired', 'property_no', 'name',  'description', 'quantity', 'condition','location')
    
    item_table_report = ttk.Treeview(sub_frame4_5, columns=columns, show='headings', height=10)
    item_table_report.pack(fill=tk.BOTH)
    
    item_table_report.column('date_acquired', width = 100)
    item_table_report.column('property_no', width = 100)
    item_table_report.column('name', width = 100)
    item_table_report.column('description', width = 100)
    item_table_report.column('quantity', width = 100)
    item_table_report.column('condition', width = 100)
    item_table_report.column('location', width = 100)
    
    item_table_report.heading('date_acquired', text='Date Acquired')
    item_table_report.heading('property_no', text='Property No.')
    item_table_report.heading('name', text='Name')
    item_table_report.heading('description', text='Description')
    item_table_report.heading('quantity', text='Quantity')
    item_table_report.heading('condition', text='Condition')
    item_table_report.heading('location', text='Location')
    
    
    #reasons
    frame4 = ctk.CTkFrame(tab2)
    frame4.pack(fill = 'both', pady = 5, padx = 10)
    
    sub_frame5 = ctk.CTkFrame(frame4, fg_color='transparent')
    sub_frame5.pack(fill = 'x')
    reason_label = ctk.CTkLabel(sub_frame5, text = 'Reasons for Transfer: ')
    reason_label.pack(side = 'left', pady = 5, padx = 10)
    
    def split_text(event=None):
        text = entry.get("1.0", "end-1c")  # Retrieve the entire text with newlines included
        lines = text.split("\n")

        line1 = ""
        line2 = ""
        line3 = ""

        for line in lines:
            if len(line1) + len(line2) + len(line3) >= 240:
                entry.configure(state="disabled")  # Disable text entry when the character limit is reached on all lines
                break

            while len(line) > 80:
                if len(line1) + len(line2) + len(line3) >= 240:
                    entry.delete("end-2c", "end-1c")  # Remove the last character if the limit is reached
                    entry.configure(state="disabled")  # Disable text entry when the character limit is reached on all lines
                    return

                line_chunk = line[:80]  # Get the first 50 characters
                line = line[80:]  # Remove the first 50 characters from the line
                if line1 == "":
                    line1 = line_chunk
                elif line2 == "":
                    line2 = line_chunk
                elif line3 == "":
                    line3 = line_chunk

            if len(line1) + len(line2) + len(line3) >= 240:
                entry.delete("end-2c", "end-1c")  # Remove the last character if the limit is reached
                entry.configure(state="disabled")  # Disable text entry when the character limit is reached on all lines
                return

            if line1 == "":
                line1 = line
            elif line2 == "":
                line2 = line
            elif line3 == "":
                line3 = line
        
        book = load_workbook('./resources/user/data/UserLog.xlsx')
        sheet2 = book['Report']
        sheet2['A1'] = line1
        sheet2['A2'] = line2
        sheet2['A3'] = line3
        book.save('./resources/user/data/UserLog.xlsx')
        
        print("Line 1:", line1)
        print("Line 2:", line2)
        print("Line 3:", line3)
        
    def handle_keypress(event):
        if len(entry.get("1.0", "end-1c")) >= 240 and event.keysym != "BackSpace":
            return "break"  # Stop further keypress events when the character limit is reached on all lines
        else:
            return None
    
    def handle_paste(event):
        if len(entry.get("1.0", "end-1c")) + len(event.widget.clipboard_get()) > 240:
            return "break"  # Prevent pasting if the character limit will be exceeded
        else:
            report_form.after(10, enable_editing)  # Enable editing after the text is pasted

    def enable_editing():
        entry.config(state="normal")  # Enable text editing 
        
    entry = ctk.CTkTextbox(frame4, height=90, fg_color='white', wrap="word")
    entry.pack(fill = 'both', pady = 10, padx = 10)# Use CtkTextbox widget with word wrap
    entry.configure(state="normal")  # Enable text editing initially

    entry.bind("<KeyPress>", handle_keypress)  # Bind keypress event to the handle_keypress function
    entry.bind("<KeyRelease>", split_text)  # Trigger split_text function on Enter key press
    entry.bind("<Control-v>", handle_paste) 

    entry.pack(fill="both", expand=True)
    
    #third tab
    #approved by
    frame5 = ctk.CTkFrame(tab3, fg_color='#F4F4F4')
    frame5.pack(fill = 'both', pady = 5, padx = 10)
    
    sub_frame6 = ctk.CTkFrame(frame5, fg_color='transparent')
    sub_frame6.pack(fill = 'x')
    
    approved_by_label = ctk.CTkLabel(sub_frame6, text = 'Approved by: ')
    approved_by_label.pack(side = 'left', pady = 3, padx = 10)
    
    sub_frame6_2 = ctk.CTkFrame(frame5, fg_color='transparent')
    sub_frame6_2.pack(fill = 'x')
    
    name_label = ctk.CTkLabel(sub_frame6_2, text = 'Printed Name: ')
    name_label.pack(side = 'left', pady = 3, padx = 10)
    approve_name = ctk.CTkEntry(sub_frame6_2, width = 300, fg_color='white')
    approve_name.pack(side = 'left', pady = 3, padx = 10)
    
    designation_label = ctk.CTkLabel(sub_frame6_2, text = 'Designation: ')
    designation_label.pack(side = 'left', pady = 3, padx = 10)
    approve_designation = ctk.CTkEntry(sub_frame6_2, width = 200, fg_color='white')
    approve_designation.pack(side = 'left', pady = 3, padx = 10)
    
    approve_date_label = ctk.CTkLabel(sub_frame6_2, text = 'Date: ')
    approve_date_label.pack(side = 'left', pady = 3, padx = 10)
    approve_date = DateEntry(sub_frame6_2, font= ctk.CTkFont('Arial', size = 12) )
    approve_date.pack(side = 'left', pady = 3, padx = 10)
    
    #Released/issued by
    frame6 = ctk.CTkFrame(tab3, fg_color='#F4F4F4')
    frame6.pack(fill = 'both', pady = 5, padx = 10) 
    
    sub_frame7 = ctk.CTkFrame(frame6, fg_color='transparent')
    sub_frame7.pack(fill = 'x')
    
    released_by_label = ctk.CTkLabel(sub_frame7, text = 'Released/Issued by: ')
    released_by_label.pack(side = 'left', pady = 3, padx = 10)
    
    sub_frame7_2 = ctk.CTkFrame(frame6, fg_color='transparent')
    sub_frame7_2.pack(fill = 'x')
    
    name_label = ctk.CTkLabel(sub_frame7_2, text = 'Printed Name: ')
    name_label.pack(side = 'left', pady = 3, padx = 10)
    release_name = ctk.CTkEntry(sub_frame7_2, width = 300, fg_color='white')
    release_name.pack(side = 'left', pady = 3, padx = 10)
    
    designation_label = ctk.CTkLabel(sub_frame7_2, text = 'Designation: ')
    designation_label.pack(side = 'left', pady = 3, padx = 10)
    release_designation = ctk.CTkEntry(sub_frame7_2, width = 200, fg_color='white')
    release_designation.pack(side = 'left', pady = 3, padx = 10)
    
    release_date_label = ctk.CTkLabel(sub_frame7_2, text = 'Date: ')
    release_date_label.pack(side = 'left', pady = 3, padx = 10)
    release_date = DateEntry(sub_frame7_2, font= ctk.CTkFont('Arial', size = 12) )
    release_date.pack(side = 'left', pady = 3, padx = 10)
    
    #receive by
    frame7 = ctk.CTkFrame(tab3, fg_color='#F4F4F4')
    frame7.pack(fill = 'both', pady = 5, padx = 10) 
    
    sub_frame8 = ctk.CTkFrame(frame7, fg_color='transparent')
    sub_frame8.pack(fill = 'x')
    
    received_by_label = ctk.CTkLabel(sub_frame8, text = 'Received by: ')
    received_by_label.pack(side = 'left', pady = 5, padx = 10)
    
    sub_frame8_2 = ctk.CTkFrame(frame7, fg_color='transparent')
    sub_frame8_2.pack(fill = 'x')
    
    name_label = ctk.CTkLabel(sub_frame8_2, text = 'Printed Name: ')
    name_label.pack(side = 'left', pady = 2, padx = 10)
    receive_name = ctk.CTkEntry(sub_frame8_2, width = 300, fg_color='white')
    receive_name.pack(side = 'left', pady = 2, padx = 10)
    
    designation_label = ctk.CTkLabel(sub_frame8_2, text = 'Designation: ')
    designation_label.pack(side = 'left', pady = 2, padx = 10)
    receive_designation = ctk.CTkEntry(sub_frame8_2, width= 200, fg_color='white')
    receive_designation.pack(side = 'left', pady = 2, padx = 10)
    
    receive_date_label = ctk.CTkLabel(sub_frame8_2, text = 'Date: ')
    receive_date_label.pack(side = 'left', pady = 5, padx = 10)
    receive_date = DateEntry(sub_frame8_2, font= ctk.CTkFont('Arial', size = 12) )
    receive_date.pack(side = 'left', pady = 2, padx = 10)
    
    #generate report
    gen_report = ctk.CTkButton(main, command=generate_report, text = 'Submit Report', font= ctk.CTkFont('Arial', size = 15, weight = "bold"),fg_color='#0077B8', text_color='white', corner_radius = 30)
    gen_report.pack(side = 'right', pady = 5, padx = (0, 20))
    
    
            
    def callback():
        from resources.py_client.client_requestlist import requestlist_window
        try:
            
            report_form.withdraw()
            requestlist_window()
            book = load_workbook('./resources/user/data/UserLog.xlsx')
            sheet2 = book['Report']
            sheet2['A1'] = ''
            sheet2['A2'] = ''
            sheet2['A3'] = ''
           
            book.save('./resources/user/data/UserLog.xlsx')

                
        except Exception as e:
            if messagebox.askok("ERROR", e ):
                print(e)
                
    refresh_treeview()
    report_form.protocol("WM_DELETE_WINDOW", callback)
    report_form.mainloop()

    
            
                
#request_window()
