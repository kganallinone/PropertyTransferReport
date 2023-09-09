import tkinter as tk
from tkinter import ttk
import customtkinter as ctk
from PIL import Image
from tkcalendar import *
from tkinter import messagebox
import pyrebase
from openpyxl import load_workbook
import datetime
from pathlib import Path
import shutil


ctk.set_appearance_mode("light")  # Modes: system (default), light, dark
current_datetime = datetime.datetime.now()
current_day = current_datetime.strftime("%Y-%m-%d")
current_time = current_datetime.strftime("%H:%M:%S")

def showptr_window(getptr):
    
    global suggestions_frame
    suggestions_frame = None
    suggestions_frame = None
    report_form = ctk.CTkToplevel()
    report_form.title("Property Transfer | Document Info")
    report_form.geometry("{0}x{1}+200+60".format(1000,560))
    report_form.resizable(False, False)
    report_form.configure(fg_color = 'white')

    #<----------------FIREBASE CONFIGURATION--------------------------------->#
    config = {
        "apiKey": "AIzaSyCyCtXg05ff-tocdUaxjBFfa550TLfKZQ4",
        "authDomain": "puplopez-ptp.firebaseapp.com",
        "projectId": "puplopez-ptp",
        "databaseURL": "https://puplopez-ptp-default-rtdb.firebaseio.com/",
        "storageBucket": "puplopez-ptp.appspot.com",
        "messagingSenderId": "638312293451",
        "appId": "1:638312293451:web:79e9405d11db0496f599ce",
        "measurementId": "G-9Z3SPKFR9V"
    }
    #<----------------FIREBASE CONNECT--------------------------------->#
    firebase = pyrebase.initialize_app(config)

    db = firebase.database()
    
    wb = load_workbook("./resources/user/data/UserLog.xlsx", data_only=True)
    sh = wb["User email"]
    email = sh["A1"].value
    print(email)

    #<-------------------------ITEM SUGGESTION------------------------------------>
    
    def approval_report(remarks):
        if remarks == 'APPROVED':

            ptrno = ptr_no.get()
            fc = fund_cluster.get()
            getdate = date.get()
            fromao = from_ao_afc.get()
            toao = to_ao_afc.get()
            one = transfer_var1.get()
            two = transfer_var2.get()
            three = transfer_var3.get()
            four = transfer_var4.get()
            other = transfer_4_entry.get()
            
            if len(ptrno) == 0:
                answer=messagebox.askokcancel("HELLO","Please add the PTR No before generating.")
                print(answer) 
            else:
                    
                def copy_and_rename_excel_file(source_file, destination_file):
                    # Get the path to the user's "Downloads" folder
                    downloads_folder = Path.home() / 'Downloads'
                    
                    # Join the "Downloads" folder path with the destination file name
                    destination_path = downloads_folder / destination_file
                    
                    shutil.copy(source_file, str(destination_path))

                    def replace_placeholders(file_path, placeholders):
                        workbook = load_workbook(file_path)
                        
                        for sheet in workbook:
                            for row in sheet.iter_rows():
                                for cell in row:
                                    if cell.value:
                                        cell_value = str(cell.value)
                                        for placeholder, new_value in placeholders.items():
                                            if placeholder in cell_value:
                                                cell_value = cell_value.replace(placeholder, str(new_value))
                                                cell.value = cell_value
                        
                                workbook.save(file_path)

                    generated_db = db.child("generated_files").order_by_child("PTR").equal_to(getptr).get()
                    for dt in generated_db.each():
                        L1 = dt.val()['REASON1']
                        L2 = dt.val()['REASON2']
                        L3 = dt.val()['REASON3']
                        A2 = dt.val()['DA1']
                        B2 = dt.val()['PN1']
                        C2 = dt.val()['N1']
                        D2 = dt.val()['DESC1']
                        E2 = dt.val()['AM1']
                        F2 = dt.val()['CON1']
                        G2 = dt.val()['LOC1']
                        A3 = dt.val()['DA2']
                        B3 = dt.val()['PN2']
                        C3 = dt.val()['N2']
                        D3 = dt.val()['DESC2']
                        E3 = dt.val()['AM2']
                        F3 = dt.val()['CON2']
                        G3 = dt.val()['LOC2']
                        A4 = dt.val()['DA3']
                        B4 = dt.val()['PN3']
                        C4 = dt.val()['N3']
                        D4 = dt.val()['DESC3']
                        E4 = dt.val()['AM3']
                        F4 = dt.val()['CON3']
                        G4 = dt.val()['LOC3']
                        A5 = dt.val()['DA4']
                        B5 = dt.val()['PN4']
                        C5 = dt.val()['N4']
                        D5 = dt.val()['DESC4']
                        E5 = dt.val()['AM4']
                        F5 = dt.val()['CON4']
                        G5 = dt.val()['LOC4']
                        A6 = dt.val()['DA5']
                        B6 = dt.val()['PN5']
                        C6 = dt.val()['N5']
                        D6 = dt.val()['DESC5']
                        E6 = dt.val()['AM5']
                        F6 = dt.val()['CON5']
                        G6 = dt.val()['LOC5']
                        A7 = dt.val()['DA6']
                        B7 = dt.val()['PN6']
                        C7 = dt.val()['N6']
                        D7 = dt.val()['DESC6']
                        E7 = dt.val()['AM6']
                        F7 = dt.val()['CON6']
                        G7 = dt.val()['LOC6']
                        A8 = dt.val()['DA7']
                        B8 = dt.val()['PN7']
                        C8 = dt.val()['N7']
                        D8 = dt.val()['DESC7']
                        E8 = dt.val()['AM7']
                        F8 = dt.val()['CON7']
                        G8 = dt.val()['LOC7']
                        A9 = dt.val()['DA8']
                        B9 = dt.val()['PN8']
                        C9 = dt.val()['N8']
                        D9 = dt.val()['DESC8']
                        E9 = dt.val()['AM8']
                        F9 = dt.val()['CON8']
                        G9 = dt.val()['LOC8']
                        A10 = dt.val()['DA9']
                        B10 = dt.val()['PN9']
                        C10 = dt.val()['N9']
                        D10 = dt.val()['DESC9']
                        E10 = dt.val()['AM9']
                        F10 = dt.val()['CON9']
                        G10 = dt.val()['LOC9']
                        A11 = dt.val()['DA10']
                        B11 = dt.val()['PN10']
                        C11 = dt.val()['N10']
                        D11 = dt.val()['DESC10']
                        E11 = dt.val()['AM10']
                        F11 = dt.val()['CON10']
                        G11 = dt.val()['LOC10']
                        A12 = dt.val()['DA11']
                        B12 = dt.val()['PN11']
                        C12 = dt.val()['N11']
                        D12 = dt.val()['DESC11']
                        E12 = dt.val()['AM11']
                        F12 = dt.val()['CON11']
                        G12 = dt.val()['LOC11']
                        A13 = dt.val()['DA12']
                        B13 = dt.val()['PN12']
                        C13 = dt.val()['N12']
                        D13 = dt.val()['DESC12']
                        E13 = dt.val()['AM12']
                        F13 = dt.val()['CON12']
                        G13 = dt.val()['LOC12']
                        A14 = dt.val()['DA13']
                        B14 = dt.val()['PN13']
                        C14 = dt.val()['N13']
                        D14 = dt.val()['DESC13']
                        E14 = dt.val()['AM13']
                        F14 = dt.val()['CON13']
                        G14 = dt.val()['LOC13']
                        A15 = dt.val()['DA14']
                        B15 = dt.val()['PN14']
                        C15 = dt.val()['N14']
                        D15 = dt.val()['DESC14']
                        E15 = dt.val()['AM14']
                        F15 = dt.val()['CON14']
                        G15 = dt.val()['LOC14']
                        A16 = dt.val()['DA15']
                        B16 = dt.val()['PN15']
                        C16 = dt.val()['N15']
                        D16 = dt.val()['DESC15']
                        E16 = dt.val()['AM15']
                        F16 = dt.val()['CON15']
                        G16 = dt.val()['LOC15']
                        A17 = dt.val()['DA16']
                        B17 = dt.val()['PN16']
                        C17 = dt.val()['N16']
                        D17 = dt.val()['DESC16']
                        E17 = dt.val()['AM16']
                        F17 = dt.val()['CON16']
                        G17 = dt.val()['LOC16']
                        A18 = dt.val()['DA17']
                        B18 = dt.val()['PN17']
                        C18 = dt.val()['N17']
                        D18 = dt.val()['DESC17']
                        E18 = dt.val()['AM17']
                        F18 = dt.val()['CON17']
                        G18 = dt.val()['LOC17']
                        CD2 = ''
                        CD3 = ''
                        CD4 = ''
                        CD5 = ''
                        CD6 = ''
                        CD7 = ''
                        CD8 = ''
                        CD9 = ''
                        CD10 = ''
                        CD11 = ''
                        CD12 = ''
                        CD13 = ''
                        CD14 = ''
                        CD15 = ''
                        CD16 = ''
                        CD17 = ''
                        CD18 = ''
                    
                    
                    #2
                    if A2 == '':
                        A2 = ''
                    else:
                        db.child("items_info").child(B2).update({"DTA":A2})
                        
                    if B2 == '':
                        B2 = ''
                    else:
                        db.child("items_info").child(B2).update({"NO":B2})
                        
                    if C2 == '' or D2 == '':
                        print(0)
                    else:
                        if C2 == '':
                            CD2 = D2
                            db.child("items_info").child(B2).update({"DES":CD2})
                        elif D2 == '':
                            CD2 = C2
                            db.child("items_info").child(B2).update({"NAME":CD2})
                        else:
                            CD2 = C2+"-"+D2
                            db.child("items_info").child(B2).update({"NAME":C2, "DES":D2})
                    if E2 == '':
                        E2 = ''
                    else:
                        db.child("items_info").child(B2).update({"QTY":E2})
                        
                    if F2 == '':
                        F2 = '' 
                    else:
                        db.child("items_info").child(B2).update({"CON":F2})
                        
                    if G2 == '':
                        G2 = '' 
                    else:
                        db.child("items_info").child(B2).update({"LOC":G2})
                    
                    #3    
                    if A3 == '':
                        A3 = ''
                    else:
                        db.child("items_info").child(B3).update({"DTA":A3})
                    
                    if B3 == '':
                        B3 = ''
                    else:
                        db.child("items_info").child(B3).update({"NO":B3})
                    
                    if C3== '' or D3== '':
                        print(0)
                    else:
                        if C3== '':
                            CD3 = D3
                            db.child("items_info").child(B3).update({"DES":CD3})
                        elif D3 == '':
                            CD3 = C3
                            db.child("items_info").child(B3).update({"NAME":CD3})
                        else:
                            CD3 = C3+"-"+D3
                            db.child("items_info").child(B3).update({"NAME":C3, "DES":D3})
                    if E3 == '':
                        E3 = ''
                    else:
                        db.child("items_info").child(B3).update({"QTY":E3})
                        
                    if F3 == '':
                        F3 = ''  
                    else:
                        db.child("items_info").child(B3).update({"CON":F3})
                        
                    if G3 == '':
                        G3 = '' 
                    else:
                        db.child("items_info").child(B3).update({"LOC":G3})
                        
                    #4       
                    if A4 == '':
                        A4 = ''
                    else:
                        db.child("items_info").child(B4).update({"DTA":A4})
                    
                    if B4 == '':
                        B4 = ''
                    else:
                        db.child("items_info").child(B4).update({"NO":B4})
                    
                    if C4 == '' or D4 == '':
                        print(0)
                    else:
                        if C4 == '':
                            CD4 = D4
                            db.child("items_info").child(B4).update({"DES":CD4})
                        elif D4 == '':
                            CD4 = C4
                            db.child("items_info").child(B4).update({"NAME":CD4})
                        else:
                            CD4 = C4+"-"+D4
                            db.child("items_info").child(B4).update({"NAME":C4, "DES":D4})
                    if E4 == '':
                        E4 = ''
                    else:
                        db.child("items_info").child(B4).update({"QTY":E4})
                        
                    if F4 == '':
                        F4 = ''  
                    else:
                        db.child("items_info").child(B4).update({"CON":F4})
                        
                    if G4 == '':
                        G4 = '' 
                    else:
                        db.child("items_info").child(B4).update({"LOC":G4})
                        
                    #5    
                    if A5 == '':
                        A5 = ''
                    else:
                        db.child("items_info").child(B5).update({"DTA":A5})
                    
                    if B5 == '':
                        B5 = ''
                    else:
                        db.child("items_info").child(B5).update({"NO":B5})
                    
                    if C5 == '' or D5 == '':
                        print(0)
                    else:
                        if C5 == '':
                            CD5 = D5
                            db.child("items_info").child(B5).update({"DES":CD5})
                        elif D5 == '':
                            CD5 = C5
                            db.child("items_info").child(B5).update({"NAME":CD5})
                        else:
                            CD5 = C5+"-"+D5
                            db.child("items_info").child(B5).update({"NAME":C5, "DES":D5})
                    if E5 == '':
                        E5 = ''
                    else:
                        db.child("items_info").child(B5).update({"QTY":E5})
                        
                    if F5 == '':
                        F5 = ''  
                    else:
                        db.child("items_info").child(B5).update({"CON":F4})
                        
                    if G5 == '':
                        G5 = '' 
                    else:
                        db.child("items_info").child(B5).update({"LOC":G5})
                    
                    #6    
                    if A6 == '':
                        A6 = ''
                    else:
                        db.child("items_info").child(B6).update({"DTA":A6})
                    
                    if B6 == '':
                        B6 = ''
                    else:
                        db.child("items_info").child(B6).update({"NO":B6})
                    
                    if C6 == '' or D6 == '':
                        print(0)
                    else:
                        if C6 == '':
                            CD6 = D6
                            db.child("items_info").child(B6).update({"DES":CD6})
                        elif D6 == '':
                            CD6 = C6
                            db.child("items_info").child(B6).update({"NAME":CD6})
                        else:
                            CD6 = C6+"-"+D6
                            db.child("items_info").child(B6).update({"NAME":C6, "DES":D6})
                    if E6 == '':
                        E6 = ''
                    else:
                        db.child("items_info").child(B6).update({"QTY":E6})
                        
                    if F6 == '':
                        F6 = ''  
                    else:
                        db.child("items_info").child(B6).update({"CON":F6})
                        
                    if G6 == '':
                        G6 = '' 
                    else:
                        db.child("items_info").child(B6).update({"LOC":G6})
                    #7    
                    
                    if A7 == '':
                        A7 = ''
                    else:
                        db.child("items_info").child(B7).update({"DTA":A7})
                    
                    if B7 == '':
                        B7 = ''
                    else:
                        db.child("items_info").child(B7).update({"NO":B7})
                    
                    if C7 == '' or D7 == '':
                        print(0)
                    else:
                        if C7 == '':
                            CD7 = D7
                            db.child("items_info").child(B7).update({"DES":CD7})
                        elif D7 == '':
                            CD7 = C7
                            db.child("items_info").child(B7).update({"NAME":CD7})
                        else:
                            CD7 = C7+"-"+D7
                            db.child("items_info").child(B7).update({"NAME":C7, "DES":D7})
                    if E7 == '':
                        E7 = ''
                    else:
                        db.child("items_info").child(B7).update({"QTY":E7})
                        
                    if F7 == '':
                        F7 = ''  
                    else:
                        db.child("items_info").child(B7).update({"CON":F7})
                        
                    if G7 == '':
                        G7 = '' 
                    else:
                        db.child("items_info").child(B7).update({"LOC":G7}) 
                    #8    
                    if A8 == '':
                        A8 = ''
                    else:
                        db.child("items_info").child(B8).update({"DTA":A8})
                    
                    if B8 == '':
                        B8 = ''
                    else:
                        db.child("items_info").child(B8).update({"NO":B8})
                    
                    if C8 == '' or D8 == '':
                        print(0)
                    else:
                        if C8 == '':
                            CD8 = D8
                            db.child("items_info").child(B8).update({"DES":CD8})
                        elif D8 == '':
                            CD8 = C8
                            db.child("items_info").child(B8).update({"NAME":CD8})
                        else:
                            CD8 = C8+"-"+D8
                            db.child("items_info").child(B8).update({"NAME":C8, "DES":D8})
                    if E8 == '':
                        E8 = ''
                    else:
                        db.child("items_info").child(B8).update({"QTY":E8})
                        
                    if F8 == '':
                        F8 = ''  
                    else:
                        db.child("items_info").child(B8).update({"CON":F8})
                        
                    if G8 == '':
                        G8 = '' 
                    else:
                        db.child("items_info").child(B8).update({"LOC":G8})   
                    #9    
                    if A9 == '':
                        A9 = ''
                    else:
                        db.child("items_info").child(B9).update({"DTA":A9})
                    
                    if B9 == '':
                        B9 = ''
                    else:
                        db.child("items_info").child(B9).update({"NO":B9})
                    
                    if C9 == '' or D9 == '':
                        print(0)
                    else:
                        if C9 == '':
                            CD9 = D9
                            db.child("items_info").child(B9).update({"DES":CD9})
                        elif D9 == '':
                            CD9 = C9
                            db.child("items_info").child(B9).update({"NAME":CD9})
                        else:
                            CD9 = C9+"-"+D9
                            db.child("items_info").child(B9).update({"NAME":C9, "DES":D9})
                    if E9 == '':
                        E9 = ''
                    else:
                        db.child("items_info").child(B9).update({"QTY":E9})
                        
                    if F9 == '':
                        F9 = ''  
                    else:
                        db.child("items_info").child(B9).update({"CON":F9})
                        
                    if G9 == '':
                        G9 = '' 
                    else:
                        db.child("items_info").child(B9).update({"LOC":G9}) 
                    #10    
                    if A10 == '':
                        A10 = ''
                    else:
                        db.child("items_info").child(B10).update({"DTA":A10})
                    
                    if B10 == '':
                        B10 = ''
                    else:
                        db.child("items_info").child(B10).update({"NO":B10})
                    
                    if C10 == '' or D10 == '':
                        print(0)
                    else:
                        if C10 == '':
                            CD10 = D10
                            db.child("items_info").child(B10).update({"DES":CD10})
                        elif D10 == '':
                            CD10 = C10
                            db.child("items_info").child(B10).update({"NAME":CD10})
                        else:
                            CD10 = C10+"-"+D10
                            db.child("items_info").child(B10).update({"NAME":C10, "DES":D10})
                    if E10 == '':
                        E10 = ''
                    else:
                        db.child("items_info").child(B10).update({"QTY":E10})
                        
                    if F10 == '':
                        F10 = ''  
                    else:
                        db.child("items_info").child(B10).update({"CON":F10})
                        
                    if G10 == '':
                        G10 = '' 
                    else:
                        db.child("items_info").child(B10).update({"LOC":G10})   
                    #11
                    if A11 == '':
                        A11 = ''
                    else:
                        db.child("items_info").child(B11).update({"DTA":A11})
                    
                    if B11 == '':
                        B11 = ''
                    else:
                        db.child("items_info").child(B11).update({"NO":B11})
                    
                    if C11 == '' or D11 == '':
                        print(0)
                    else:
                        if C11 == '':
                            CD11 = D11
                            db.child("items_info").child(B11).update({"DES":CD11})
                        elif D11 == '':
                            CD11 = C11
                            db.child("items_info").child(B11).update({"NAME":CD11})
                        else:
                            CD11 = C11+"-"+D11
                            db.child("items_info").child(B11).update({"NAME":C11, "DES":D11})
                    if E11 == '':
                        E11 = ''
                    else:
                        db.child("items_info").child(B11).update({"QTY":E11})
                        
                    if F11 == '':
                        F11 = ''  
                    else:
                        db.child("items_info").child(B11).update({"CON":F11})
                        
                    if G11 == '':
                        G11 = '' 
                    else:
                        db.child("items_info").child(B11).update({"LOC":G11})
                    #12
                    if A12 == '':
                        A12 = ''
                    else:
                        db.child("items_info").child(B12).update({"DTA":A12})
                        
                    if B12 == '':
                        B12 = ''
                    else:
                        db.child("items_info").child(B12).update({"NO":B12})
                        
                    if C12== '' or D12== '':
                        print(0)
                    else:
                        if C12 == '':
                            CD12 = D12
                            db.child("items_info").child(B12).update({"DES":CD12})
                        elif D12 == '':
                            CD12 = C12
                            db.child("items_info").child(B12).update({"NAME":CD12})
                        else:
                            CD12 = C12+"-"+D12
                            db.child("items_info").child(B2).update({"NAME":C12, "DES":D12})
                    if E12 == '':
                        E12 = ''
                    else:
                        db.child("items_info").child(B12).update({"QTY":E12})
                        
                    if F12 == '':
                        F12 = '' 
                    else:
                        db.child("items_info").child(B12).update({"CON":F12})
                        
                    if G12 == '':
                        G12 = '' 
                    else:
                        db.child("items_info").child(B12).update({"LOC":G12})
                    
                    #13    
                    if A13 == '':
                        A13 = ''
                    else:
                        db.child("items_info").child(B13).update({"DTA":A13})
                    
                    if B13 == '':
                        B13 = ''
                    else:
                        db.child("items_info").child(B13).update({"NO":B13})
                    
                    if C13== '' or D13== '':
                        print(0)
                    else:
                        if C13== '':
                            CD13 = D13
                            db.child("items_info").child(B13).update({"DES":CD13})
                        elif D13 == '':
                            CD13 = C13
                            db.child("items_info").child(B13).update({"NAME":CD13})
                        else:
                            CD13 = C13+"-"+D13
                            db.child("items_info").child(B13).update({"NAME":C13, "DES":D13})
                    if E13 == '':
                        E13 = ''
                    else:
                        db.child("items_info").child(B13).update({"QTY":E13})
                        
                    if F13 == '':
                        F13 = ''  
                    else:
                        db.child("items_info").child(B13).update({"CON":F13})
                        
                    if G13 == '':
                        G13 = '' 
                    else:
                        db.child("items_info").child(B13).update({"LOC":G13})
                        
                    #14       
                    if A14 == '':
                        A14 = ''
                    else:
                        db.child("items_info").child(B14).update({"DTA":A14})
                    
                    if B14 == '':
                        B14 = ''
                    else:
                        db.child("items_info").child(B14).update({"NO":B14})
                    
                    if C14 == '' or D14 == '':
                        print(0)
                    else:
                        if C14 == '':
                            CD14 = D14
                            db.child("items_info").child(B14).update({"DES":CD14})
                        elif D14 == '':
                            CD14 = C14
                            db.child("items_info").child(B14).update({"NAME":CD14})
                        else:
                            CD14 = C14+"-"+D14
                            db.child("items_info").child(B14).update({"NAME":C14, "DES":D14})
                    if E14 == '':
                        E14 = ''
                    else:
                        db.child("items_info").child(B14).update({"QTY":E14})
                        
                    if F14 == '':
                        F14 = ''  
                    else:
                        db.child("items_info").child(B14).update({"CON":F14})
                        
                    if G14 == '':
                        G14 = '' 
                    else:
                        db.child("items_info").child(B14).update({"LOC":G14})
                        
                    #15    
                    if A15 == '':
                        A15 = ''
                    else:
                        db.child("items_info").child(B15).update({"DTA":A15})
                    
                    if B15 == '':
                        B15 = ''
                    else:
                        db.child("items_info").child(B15).update({"NO":B15})
                    
                    if C15 == '' or D15 == '':
                        print(0)
                    else:
                        if C15 == '':
                            CD15 = D15
                            db.child("items_info").child(B15).update({"DES":CD15})
                        elif D15 == '':
                            CD15 = C15
                            db.child("items_info").child(B15).update({"NAME":CD15})
                        else:
                            CD15 = C15+"-"+D15
                            db.child("items_info").child(B15).update({"NAME":C15, "DES":D15})
                    if E15 == '':
                        E15 = ''
                    else:
                        db.child("items_info").child(B15).update({"QTY":E15})
                        
                    if F15 == '':
                        F15 = ''  
                    else:
                        db.child("items_info").child(B15).update({"CON":F15})
                        
                    if G15 == '':
                        G15 = '' 
                    else:
                        db.child("items_info").child(B15).update({"LOC":G15})
                    
                    #16    
                    if A16 == '':
                        A16 = ''
                    else:
                        db.child("items_info").child(B16).update({"DTA":A16})
                    
                    if B16 == '':
                        B16 = ''
                    else:
                        db.child("items_info").child(B16).update({"NO":B16})
                    
                    if C16 == '' or D16 == '':
                        print(0)
                    else:
                        if C16 == '':
                            CD16 = D16
                            db.child("items_info").child(B16).update({"DES":CD16})
                        elif D16 == '':
                            CD16 = C16
                            db.child("items_info").child(B16).update({"NAME":CD16})
                        else:
                            CD6 = C6+"-"+D6
                            db.child("items_info").child(B16).update({"NAME":C16, "DES":D16})
                    if E16 == '':
                        E16 = ''
                    else:
                        db.child("items_info").child(B16).update({"QTY":E16})
                        
                    if F16 == '':
                        F16 = ''  
                    else:
                        db.child("items_info").child(B16).update({"CON":F16})
                        
                    if G16 == '':
                        G16 = '' 
                    else:
                        db.child("items_info").child(B16).update({"LOC":G16})
                    #17    
                    
                    if A17 == '':
                        A17 = ''
                    else:
                        db.child("items_info").child(B17).update({"DTA":A17})
                    
                    if B17 == '':
                        B17 = ''
                    else:
                        db.child("items_info").child(B17).update({"NO":B17})
                    
                    if C17 == '' or D17 == '':
                        print(0)
                    else:
                        if C17 == '':
                            CD17 = D17
                            db.child("items_info").child(B17).update({"DES":CD17})
                        elif D7 == '':
                            CD17 = C17
                            db.child("items_info").child(B17).update({"NAME":CD17})
                        else:
                            CD17 = C17+"-"+D17
                            db.child("items_info").child(B17).update({"NAME":C17, "DES":D17})
                    if E17 == '':
                        E17 = ''
                    else:
                        db.child("items_info").child(B17).update({"QTY":E17})
                        
                    if F17 == '':
                        F17 = ''  
                    else:
                        db.child("items_info").child(B17).update({"CON":F17})
                        
                    if G17 == '':
                        G17 = '' 
                    else:
                        db.child("items_info").child(B17).update({"LOC":G17}) 
                    #8    
                    if A18 == '':
                        A18 = ''
                    else:
                        db.child("items_info").child(B18).update({"DTA":A18})
                    
                    if B18 == '':
                        B18 = ''
                    else:
                        db.child("items_info").child(B18).update({"NO":B18})
                    
                    if C18 == '' or D18 == '':
                        print(0)
                    else:
                        if C18 == '':
                            CD18 = D18
                            db.child("items_info").child(B18).update({"DES":CD18})
                        elif D8 == '':
                            CD18 = C18
                            db.child("items_info").child(B18).update({"NAME":CD18})
                        else:
                            CD18 = C18+"-"+D18
                            db.child("items_info").child(B18).update({"NAME":C18, "DES":D18})
                    if E18 == '':
                        E18 = ''
                    else:
                        db.child("items_info").child(B18).update({"QTY":E18})
                        
                    if F18 == '':
                        F18 = ''  
                    else:
                        db.child("items_info").child(B18).update({"CON":F18})
                        
                    if G18 == '':
                        G18 = '' 
                    else:
                        db.child("items_info").child(B18).update({"LOC":G18}) 
                                            
                    # Example usage
                    file_path = destination_path
                    if one == "on":
                        placeholders = {
                            '{{PTR}}': ptrno,
                            '{{FROM}}': fromao,
                            '{{TO}}': toao,
                            '{{FC}}': fc,
                            '{{DT}}': getdate,
                            '{{OTHER}}': '_________',
                            '{{1}}':'✔',
                            '{{2}}':'',
                            '{{3}}':'',
                            '{{4}}':'',
                            '{{REASON1}}':L1,
                            '{{REASON2}}': L2,
                            '{{REASON3}}': L3,
                            '{{DA1}}': A2,
                            '{{PN1}}': B2,
                            '{{DESC1}}': CD2,
                            '{{AM1}}': E2,
                            '{{CON1}}': F2,
                            '{{DA2}}': A3,
                            '{{PN2}}': B3,
                            '{{DESC2}}': CD3,
                            '{{AM2}}': E3,
                            '{{CON2}}': F3,
                            '{{DA3}}': A4,
                            '{{PN3}}': B4,
                            '{{DESC3}}': CD4,
                            '{{AM3}}': E4,
                            '{{CON3}}': F4,
                            '{{DA4}}': A5,
                            '{{PN4}}': B5,
                            '{{DESC4}}': CD5,
                            '{{AM4}}': E5,
                            '{{CON4}}': F5,
                            '{{DA5}}': A6,
                            '{{PN5}}': B6,
                            '{{DESC5}}': CD6,
                            '{{AM5}}': E6,
                            '{{CON5}}': F6,
                            '{{DA6}}': A7,
                            '{{PN6}}': B7,
                            '{{DESC6}}': CD7,
                            '{{AM6}}': E7,
                            '{{CON6}}': F7,
                            '{{DA7}}': A8,
                            '{{PN7}}': B8,
                            '{{DESC7}}': CD8,
                            '{{AM7}}': E8,
                            '{{CON7}}': F8,
                            '{{DA8}}': A9,
                            '{{PN8}}': B9,
                            '{{DESC8}}': CD9,
                            '{{AM8}}': E9,
                            '{{CON8}}': F9,
                            '{{DA9}}': A10,
                            '{{PN9}}': B10,
                            '{{DESC9}}': CD10,
                            '{{AM9}}': E10,
                            '{{CON9}}': F10,
                            '{{DA10}}': A11,
                            '{{PN10}}': B11,
                            '{{DESC10}}': CD11,
                            '{{AM10}}': E11,
                            '{{CON10}}': F11,
                            '{{DA11}}': A12,
                            '{{PN11}}': B12,
                            '{{DESC11}}': CD12,
                            '{{AM11}}': E12,
                            '{{CON11}}': F12,
                            '{{DA12}}': A13,
                            '{{PN12}}': B13,
                            '{{DESC12}}': CD13,
                            '{{AM12}}': E13,
                            '{{CON12}}': F13,
                            '{{DA13}}': A14,
                            '{{PN13}}': B14,
                            '{{DESC13}}': CD14,
                            '{{AM13}}': E14,
                            '{{CON13}}': F14,
                            '{{DA14}}': A15,
                            '{{PN14}}': B15,
                            '{{DESC14}}': CD15,
                            '{{AM14}}': E15,
                            '{{CON14}}': F15,
                            '{{DA15}}': A16,
                            '{{PN15}}': B16,
                            '{{DESC15}}': CD16,
                            '{{AM15}}': E16,
                            '{{CON15}}': F16,
                            '{{DA16}}': A17,
                            '{{PN16}}': B17,
                            '{{DESC16}}': CD17,
                            '{{AM16}}': E17,
                            '{{CON16}}': F17,
                            '{{DA17}}': A18,
                            '{{PN17}}': B18,
                            '{{DESC17}}': CD18,
                            '{{AM17}}': E18,
                            '{{CON17}}': F18,
                            '{{NAME}}': approve_name.get(),
                            '{{RINAME}}':release_name.get(),
                            '{{RNAME}}': receive_name.get(),
                            '{{DES}}': approve_designation.get(),
                            '{{RIDES}}':release_designation.get(),
                            '{{RDES}}':receive_designation.get(),
                            '{{ADT}}': approve_date.get(),
                            '{{RIDT}}': release_date.get(),
                            '{{RDT}}':receive_date.get()
                        }
                           
                        replace_placeholders(file_path, placeholders)
                        answer=messagebox.askokcancel("HELLO","Your PTR is successfully generated, Please document in Download.")
                        print(answer)
                        
                    if two == "on":
                        placeholders = {
                            '{{PTR}}': ptrno,
                            '{{FROM}}': fromao,
                            '{{TO}}': toao,
                            '{{FC}}': fc,
                            '{{DT}}': getdate,
                            '{{OTHER}}': '_________',
                            '{{1}}':'',
                            '{{2}}':'✔',
                            '{{3}}':'',
                            '{{4}}':'',
                            '{{REASON1}}':L1,
                            '{{REASON2}}': L2,
                            '{{REASON3}}': L3,
                            '{{DA1}}': A2,
                            '{{PN1}}': B2,
                            '{{DESC1}}': CD2,
                            '{{AM1}}': E2,
                            '{{CON1}}': F2,
                            '{{DA2}}': A3,
                            '{{PN2}}': B3,
                            '{{DESC2}}': CD3,
                            '{{AM2}}': E3,
                            '{{CON2}}': F3,
                            '{{DA3}}': A4,
                            '{{PN3}}': B4,
                            '{{DESC3}}': CD4,
                            '{{AM3}}': E4,
                            '{{CON3}}': F4,
                            '{{DA4}}': A5,
                            '{{PN4}}': B5,
                            '{{DESC4}}': CD5,
                            '{{AM4}}': E5,
                            '{{CON4}}': F5,
                            '{{DA5}}': A6,
                            '{{PN5}}': B6,
                            '{{DESC5}}': CD6,
                            '{{AM5}}': E6,
                            '{{CON5}}': F6,
                            '{{DA6}}': A7,
                            '{{PN6}}': B7,
                            '{{DESC6}}': CD7,
                            '{{AM6}}': E7,
                            '{{CON6}}': F7,
                            '{{DA7}}': A8,
                            '{{PN7}}': B8,
                            '{{DESC7}}': CD8,
                            '{{AM7}}': E8,
                            '{{CON7}}': F8,
                            '{{DA8}}': A9,
                            '{{PN8}}': B9,
                            '{{DESC8}}': CD9,
                            '{{AM8}}': E9,
                            '{{CON8}}': F9,
                            '{{DA9}}': A10,
                            '{{PN9}}': B10,
                            '{{DESC9}}': CD10,
                            '{{AM9}}': E10,
                            '{{CON9}}': F10,
                            '{{DA10}}': A11,
                            '{{PN10}}': B11,
                            '{{DESC10}}': CD11,
                            '{{AM10}}': E11,
                            '{{CON10}}': F11,
                            '{{DA11}}': A12,
                            '{{PN11}}': B12,
                            '{{DESC11}}': CD12,
                            '{{AM11}}': E12,
                            '{{CON11}}': F12,
                            '{{DA12}}': A13,
                            '{{PN12}}': B13,
                            '{{DESC12}}': CD13,
                            '{{AM12}}': E13,
                            '{{CON12}}': F13,
                            '{{DA13}}': A14,
                            '{{PN13}}': B14,
                            '{{DESC13}}': CD14,
                            '{{AM13}}': E14,
                            '{{CON13}}': F14,
                            '{{DA14}}': A15,
                            '{{PN14}}': B15,
                            '{{DESC14}}': CD15,
                            '{{AM14}}': E15,
                            '{{CON14}}': F15,
                            '{{DA15}}': A16,
                            '{{PN15}}': B16,
                            '{{DESC15}}': CD16,
                            '{{AM15}}': E16,
                            '{{CON15}}': F16,
                            '{{DA16}}': A17,
                            '{{PN16}}': B17,
                            '{{DESC16}}': CD17,
                            '{{AM16}}': E17,
                            '{{CON16}}': F17,
                            '{{DA17}}': A18,
                            '{{PN17}}': B18,
                            '{{DESC17}}': CD18,
                            '{{AM17}}': E18,
                            '{{CON17}}': F18,
                            '{{NAME}}': approve_name.get(),
                            '{{RINAME}}':release_name.get(),
                            '{{RNAME}}': receive_name.get(),
                            '{{DES}}': approve_designation.get(),
                            '{{RIDES}}':release_designation.get(),
                            '{{RDES}}':receive_designation.get(),
                            '{{ADT}}': approve_date.get(),
                            '{{RIDT}}': release_date.get(),
                            '{{RDT}}':receive_date.get()
                        }
                    
                        replace_placeholders(file_path, placeholders)
                        answer=messagebox.askokcancel("HELLO","Your PTR is successfully generated, Please document in Download.")    

                    if three == "on":
                        placeholders = {
                            '{{PTR}}': ptrno,
                            '{{FROM}}': fromao,
                            '{{TO}}': toao,
                            '{{FC}}': fc,
                            '{{DT}}': getdate,
                            '{{OTHER}}': '_________',
                            '{{1}}':'',
                            '{{2}}':'',
                            '{{3}}':'✔',
                            '{{4}}':'',
                            '{{REASON1}}':L1,
                            '{{REASON2}}': L2,
                            '{{REASON3}}': L3,
                            '{{DA1}}': A2,
                            '{{PN1}}': B2,
                            '{{DESC1}}': CD2,
                            '{{AM1}}': E2,
                            '{{CON1}}': F2,
                            '{{DA2}}': A3,
                            '{{PN2}}': B3,
                            '{{DESC2}}': CD3,
                            '{{AM2}}': E3,
                            '{{CON2}}': F3,
                            '{{DA3}}': A4,
                            '{{PN3}}': B4,
                            '{{DESC3}}': CD4,
                            '{{AM3}}': E4,
                            '{{CON3}}': F4,
                            '{{DA4}}': A5,
                            '{{PN4}}': B5,
                            '{{DESC4}}': CD5,
                            '{{AM4}}': E5,
                            '{{CON4}}': F5,
                            '{{DA5}}': A6,
                            '{{PN5}}': B6,
                            '{{DESC5}}': CD6,
                            '{{AM5}}': E6,
                            '{{CON5}}': F6,
                            '{{DA6}}': A7,
                            '{{PN6}}': B7,
                            '{{DESC6}}': CD7,
                            '{{AM6}}': E7,
                            '{{CON6}}': F7,
                            '{{DA7}}': A8,
                            '{{PN7}}': B8,
                            '{{DESC7}}': CD8,
                            '{{AM7}}': E8,
                            '{{CON7}}': F8,
                            '{{DA8}}': A9,
                            '{{PN8}}': B9,
                            '{{DESC8}}': CD9,
                            '{{AM8}}': E9,
                            '{{CON8}}': F9,
                            '{{DA9}}': A10,
                            '{{PN9}}': B10,
                            '{{DESC9}}': CD10,
                            '{{AM9}}': E10,
                            '{{CON9}}': F10,
                            '{{DA10}}': A11,
                            '{{PN10}}': B11,
                            '{{DESC10}}': CD11,
                            '{{AM10}}': E11,
                            '{{CON10}}': F11,
                            '{{DA11}}': A12,
                            '{{PN11}}': B12,
                            '{{DESC11}}': CD12,
                            '{{AM11}}': E12,
                            '{{CON11}}': F12,
                            '{{DA12}}': A13,
                            '{{PN12}}': B13,
                            '{{DESC12}}': CD13,
                            '{{AM12}}': E13,
                            '{{CON12}}': F13,
                            '{{DA13}}': A14,
                            '{{PN13}}': B14,
                            '{{DESC13}}': CD14,
                            '{{AM13}}': E14,
                            '{{CON13}}': F14,
                            '{{DA14}}': A15,
                            '{{PN14}}': B15,
                            '{{DESC14}}': CD15,
                            '{{AM14}}': E15,
                            '{{CON14}}': F15,
                            '{{DA15}}': A16,
                            '{{PN15}}': B16,
                            '{{DESC15}}': CD16,
                            '{{AM15}}': E16,
                            '{{CON15}}': F16,
                            '{{DA16}}': A17,
                            '{{PN16}}': B17,
                            '{{DESC16}}': CD17,
                            '{{AM16}}': E17,
                            '{{CON16}}': F17,
                            '{{DA17}}': A18,
                            '{{PN17}}': B18,
                            '{{DESC17}}': CD18,
                            '{{AM17}}': E18,
                            '{{CON17}}': F18,
                            '{{NAME}}': approve_name.get(),
                            '{{RINAME}}':release_name.get(),
                            '{{RNAME}}': receive_name.get(),
                            '{{DES}}': approve_designation.get(),
                            '{{RIDES}}':release_designation.get(),
                            '{{RDES}}':receive_designation.get(),
                            '{{ADT}}': approve_date.get(),
                            '{{RIDT}}': release_date.get(),
                            '{{RDT}}':receive_date.get()
                        }
                        
                        
                        replace_placeholders(file_path, placeholders)
                        answer=messagebox.askokcancel("HELLO","Your PTR is successfully generated, Please document in Download.")
                        
                    
                    if four == "on":
                        placeholders = {
                            '{{PTR}}': ptrno,
                            '{{FROM}}': fromao,
                            '{{TO}}': toao,
                            '{{FC}}': fc,
                            '{{DT}}': getdate,
                            '{{OTHER}}': other,
                            '{{1}}':'',
                            '{{2}}':'',
                            '{{3}}':'',
                            '{{4}}':'✔',
                            '{{REASON1}}':L1,
                            '{{REASON2}}': L2,
                            '{{REASON3}}': L3,
                            '{{DA1}}': A2,
                            '{{PN1}}': B2,
                            '{{DESC1}}': CD2,
                            '{{AM1}}': E2,
                            '{{CON1}}': F2,
                            '{{DA2}}': A3,
                            '{{PN2}}': B3,
                            '{{DESC2}}': CD3,
                            '{{AM2}}': E3,
                            '{{CON2}}': F3,
                            '{{DA3}}': A4,
                            '{{PN3}}': B4,
                            '{{DESC3}}': CD4,
                            '{{AM3}}': E4,
                            '{{CON3}}': F4,
                            '{{DA4}}': A5,
                            '{{PN4}}': B5,
                            '{{DESC4}}': CD5,
                            '{{AM4}}': E5,
                            '{{CON4}}': F5,
                            '{{DA5}}': A6,
                            '{{PN5}}': B6,
                            '{{DESC5}}': CD6,
                            '{{AM5}}': E6,
                            '{{CON5}}': F6,
                            '{{DA6}}': A7,
                            '{{PN6}}': B7,
                            '{{DESC6}}': CD7,
                            '{{AM6}}': E7,
                            '{{CON6}}': F7,
                            '{{DA7}}': A8,
                            '{{PN7}}': B8,
                            '{{DESC7}}': CD8,
                            '{{AM7}}': E8,
                            '{{CON7}}': F8,
                            '{{DA8}}': A9,
                            '{{PN8}}': B9,
                            '{{DESC8}}': CD9,
                            '{{AM8}}': E9,
                            '{{CON8}}': F9,
                            '{{DA9}}': A10,
                            '{{PN9}}': B10,
                            '{{DESC9}}': CD10,
                            '{{AM9}}': E10,
                            '{{CON9}}': F10,
                            '{{DA10}}': A11,
                            '{{PN10}}': B11,
                            '{{DESC10}}': CD11,
                            '{{AM10}}': E11,
                            '{{CON10}}': F11,
                            '{{DA11}}': A12,
                            '{{PN11}}': B12,
                            '{{DESC11}}': CD12,
                            '{{AM11}}': E12,
                            '{{CON11}}': F12,
                            '{{DA12}}': A13,
                            '{{PN12}}': B13,
                            '{{DESC12}}': CD13,
                            '{{AM12}}': E13,
                            '{{CON12}}': F13,
                            '{{DA13}}': A14,
                            '{{PN13}}': B14,
                            '{{DESC13}}': CD14,
                            '{{AM13}}': E14,
                            '{{CON13}}': F14,
                            '{{DA14}}': A15,
                            '{{PN14}}': B15,
                            '{{DESC14}}': CD15,
                            '{{AM14}}': E15,
                            '{{CON14}}': F15,
                            '{{DA15}}': A16,
                            '{{PN15}}': B16,
                            '{{DESC15}}': CD16,
                            '{{AM15}}': E16,
                            '{{CON15}}': F16,
                            '{{DA16}}': A17,
                            '{{PN16}}': B17,
                            '{{DESC16}}': CD17,
                            '{{AM16}}': E17,
                            '{{CON16}}': F17,
                            '{{DA17}}': A18,
                            '{{PN17}}': B18,
                            '{{DESC17}}': CD18,
                            '{{AM17}}': E18,
                            '{{CON17}}': F18,
                            '{{NAME}}': approve_name.get(),
                            '{{RINAME}}':release_name.get(),
                            '{{RNAME}}': receive_name.get(),
                            '{{DES}}': approve_designation.get(),
                            '{{RIDES}}':release_designation.get(),
                            '{{RDES}}':receive_designation.get(),
                            '{{ADT}}': approve_date.get(),
                            '{{RIDT}}': release_date.get(),
                            '{{RDT}}':receive_date.get()
                        }
                        
                        replace_placeholders(file_path, placeholders)
                        answer=messagebox.askokcancel("HELLO","Your PTR is successfully generated, Please document in Download.")
                        
                    
                
            # Example usage
            source_file = 'resources/user/data/Property-Transfer-Report.xlsx'
            destination_file = 'PTR-'+ptrno+'.xlsx'

            copy_and_rename_excel_file(source_file, destination_file)
        else:
            db.child("generated_files").child(getptr).remove()
            
            
    
    main = ctk.CTkFrame(report_form)
    main.pack(fill = 'both')
    #header frame
    header_frame = ctk.CTkFrame(main, fg_color='#313131', height=50, width=report_form.winfo_width(), corner_radius=0)
    header_frame.pack(fill = 'both', side='top')

    #app logo
    app_logo = ctk.CTkImage(Image.open('resources\images\DEFAULT\PTRLogo_White.png'), size=(25, 25))
    app_logo_label = ctk.CTkLabel(header_frame, image= app_logo, text=None)
    app_logo_label.pack(side = 'left', pady = 10, padx = 10)

    #app name
    app_name1 = ctk.CTkLabel(header_frame, text='PROPERTY', font= ctk.CTkFont('Arial', size = 22, weight = "bold"), text_color='white')
    app_name1.pack(side = 'left', pady = 10, padx = 10)
    app_name2 = ctk.CTkLabel(header_frame, text='TRANSFER', font= ctk.CTkFont('Arial', size = 22, weight = "bold"), text_color='#ee4444')
    app_name2.pack(side = 'left', pady = 10)

    label = ctk.CTkLabel(header_frame, text = 'Reports', font= ctk.CTkFont('Arial', size = 25, weight = "bold"), text_color='white')
    label.pack(side = 'right',  pady = 10, padx = 10)
        
    #notebook
    notebook = ttk.Notebook(main)
    notebook.pack(expand=True)
    
    # create frames
    tab1 = ctk.CTkFrame(notebook, width=1000, height=500, fg_color='white')
    tab2 = ctk.CTkFrame(notebook, width=1000, height=500, fg_color='white')
    tab3 = ctk.CTkFrame(notebook, width=1000, height=500, fg_color='white')

    tab1.pack(fill='both', expand=True)
    tab2.pack(fill='both', expand=True)
    tab3.pack(fill ='both', expand =True)

    # add frames to notebook
    notebook.add(tab1, text='PTR Information')
    notebook.add(tab2, text='Items')
    notebook.add(tab3, text='Approval')
    
    #first tab
    info_frame = ctk.CTkFrame(tab1, fg_color='white')
    info_frame.pack(fill = 'x', pady = 5, padx = 5)
    
    sub_info_frame = ctk.CTkFrame(info_frame, height=50 , corner_radius=20, fg_color='#F4F4F4')
    sub_info_frame.pack(fill = 'x', pady = 5, padx = 5)
    
    generated_db = db.child("generated_files").order_by_child("PTR").equal_to(getptr).get()
    for dt in generated_db.each():
        db_ptr = dt.val()['PTR']
        db_email = dt.val()['EMAIL']
        db_fc = dt.val()['FC']
        db_type = dt.val()['TRANSFER']
        db_created = dt.val()['CREATED']
        db_remarks = dt.val()['REMARKS']
        db_from = dt.val()['FROM'].title()
        db_to = dt.val()['TO'].title()
        db_dt = dt.val()['DT']
        reason = dt.val()['REASON1']+dt.val()['REASON2']+dt.val()['REASON3']
        db_riname = dt.val()['RINAME'].title()
        db_ridt = dt.val()['RIDT']
        db_rides = dt.val()['RIDES']
        db_rname = dt.val()['RNAME'].title()
        db_rdt = dt.val()['RDT']
        db_rdes = dt.val()['RDES']
        db_name = dt.val()['NAME'].title()
        db_adt = dt.val()['ADT']
        db_des = dt.val()['DES']
        
        items = [
        (dt.val()['DA1'], dt.val()['PN1'], dt.val()['N1'], dt.val()['DESC1'], dt.val()['AM1'], dt.val()['CON1'], dt.val()['LOC1']),
        (dt.val()['DA2'], dt.val()['PN2'], dt.val()['N2'], dt.val()['DESC2'], dt.val()['AM2'], dt.val()['CON2'], dt.val()['LOC2']),
        (dt.val()['DA3'], dt.val()['PN3'], dt.val()['N3'], dt.val()['DESC3'], dt.val()['AM3'], dt.val()['CON3'], dt.val()['LOC3']),
        (dt.val()['DA4'], dt.val()['PN4'], dt.val()['N4'], dt.val()['DESC4'], dt.val()['AM4'], dt.val()['CON4'], dt.val()['LOC4']),
        (dt.val()['DA5'], dt.val()['PN5'], dt.val()['N5'], dt.val()['DESC5'], dt.val()['AM5'], dt.val()['CON5'], dt.val()['LOC5']),
        (dt.val()['DA6'], dt.val()['PN6'], dt.val()['N6'], dt.val()['DESC6'], dt.val()['AM6'], dt.val()['CON6'], dt.val()['LOC6']),
        (dt.val()['DA7'], dt.val()['PN7'], dt.val()['N7'], dt.val()['DESC7'], dt.val()['AM7'], dt.val()['CON7'], dt.val()['LOC7']),
        (dt.val()['DA8'], dt.val()['PN8'], dt.val()['N8'], dt.val()['DESC8'], dt.val()['AM8'], dt.val()['CON8'], dt.val()['LOC8']),
        (dt.val()['DA9'], dt.val()['PN9'], dt.val()['N9'], dt.val()['DESC9'], dt.val()['AM9'], dt.val()['CON9'], dt.val()['LOC9']),
        (dt.val()['DA10'], dt.val()['PN10'], dt.val()['N10'], dt.val()['DESC10'], dt.val()['AM10'], dt.val()['CON10'], dt.val()['LOC10']),
        (dt.val()['DA11'], dt.val()['PN11'], dt.val()['N11'], dt.val()['DESC11'], dt.val()['AM11'], dt.val()['CON11'], dt.val()['LOC11']),
        (dt.val()['DA12'], dt.val()['PN12'], dt.val()['N12'], dt.val()['DESC12'], dt.val()['AM12'], dt.val()['CON12'], dt.val()['LOC12']),
        (dt.val()['DA13'], dt.val()['PN13'], dt.val()['N13'], dt.val()['DESC13'], dt.val()['AM13'], dt.val()['CON13'], dt.val()['LOC13']),
        (dt.val()['DA14'], dt.val()['PN14'], dt.val()['N14'], dt.val()['DESC14'], dt.val()['AM14'], dt.val()['CON14'], dt.val()['LOC14']),
        (dt.val()['DA15'], dt.val()['PN15'], dt.val()['N15'], dt.val()['DESC15'], dt.val()['AM15'], dt.val()['CON15'], dt.val()['LOC15']),
        (dt.val()['DA16'], dt.val()['PN16'], dt.val()['N16'], dt.val()['DESC16'], dt.val()['AM16'], dt.val()['CON16'], dt.val()['LOC16']),
        (dt.val()['DA17'], dt.val()['PN17'], dt.val()['N17'], dt.val()['DESC17'], dt.val()['AM17'], dt.val()['CON17'], dt.val()['LOC17'])
        
        ]
        
    user_db = db.child("user_accounts").order_by_child("EMAIL").equal_to(db_email).get()
    for dt in user_db.each():
        db_id = dt.val()['ID']
        db_fn = dt.val()['FN'].title()
        db_ln = dt.val()['LN'].title()    
        
    label_type = ctk.CTkLabel(sub_info_frame, text=db_type, text_color='black', font=('Arial', 22, 'bold'))
    label_type.pack(side='left',padx=(10, 5), pady=(3,3),ipadx=0, ipady=0, anchor='w')    
    label_key = ctk.CTkLabel(sub_info_frame, text=" | PTR No: "+db_ptr+" | Received by: "+db_riname, text_color='black', font=('Arial', 22))
    label_key.pack(side='left', padx=(3,0), pady=(0,0),ipadx=0, ipady=0, anchor='w')
    
    sub_info2_frame = ctk.CTkFrame(info_frame, height=50 , corner_radius=20, fg_color='#F4F4F4')
    sub_info2_frame.pack(fill = 'x', pady = 5, padx = 5)
    label_name = ctk.CTkLabel(sub_info2_frame, text="Requested by: "+db_ln+", "+db_fn+" (YOU) ● STATUS: "+db_remarks, text_color='black', font=('Arial', 13))
    label_name.pack(side='left',padx=(10, 0), pady=(0,0),ipadx=0, ipady=0, anchor='w')
    
    frame = ctk.CTkFrame(tab1, height=400, corner_radius=20, fg_color='#F4F4F4')
    frame.pack(fill = 'both', pady = 5, padx = 10)
    
    sub_frame = ctk.CTkFrame(frame, fg_color='transparent')
    sub_frame.pack(fill = 'x', pady = 5, padx = 5)
    
    entity_label = ctk.CTkLabel(sub_frame, text='Entity Name: ')
    entity_label.pack(side = 'left', pady = 5, padx = 10)
    
    entities = ["Polytechnic University of the Philippines - Lopez Branch"]
    entity_name = ctk.CTkOptionMenu(sub_frame, width=300, values=entities, fg_color= 'white', text_color='black', dropdown_fg_color='white', font= ctk.CTkFont('Arial', size = 12), dropdown_font=ctk.CTkFont('Arial', size = 12))
    entity_name.pack(side = 'left', pady = 5, padx = 10)
    
    fund_cluster_label = ctk.CTkLabel(sub_frame, text='Fund Cluster: ')
    fund_cluster_label.pack(side = 'left', pady = 5, padx = (170,0))
    fund_cluster= ctk.CTkEntry(sub_frame, fg_color='white', state="readonly")
    fund_cluster.pack(side = 'left', pady = 5, padx = 10)
    fund_cluster.configure(state="normal")  # Set the entry to normal state
    fund_cluster.delete(0, ctk.END)  # Clear the existing value
    fund_cluster.insert(0, db_fc)  # Set the new value
    fund_cluster.configure(state="readonly")  # Set the entry back to readonly state
    
    sub_frame2 = ctk.CTkFrame(frame, fg_color='transparent')
    sub_frame2.pack(fill = 'x', pady = 5, padx = 10)
    
    from_label = ctk.CTkLabel(sub_frame2, text='From Accountable Officer/Agency Cluster Fund: ')
    from_label.pack(side = 'left', pady = 5, padx = 10)
    from_ao_afc = ctk.CTkEntry(sub_frame2, width = 300, fg_color='white', state="readonly")
    from_ao_afc.pack(side = 'left', pady = 5, padx = 10)
    from_ao_afc.configure(state="normal")  # Set the entry to normal state
    from_ao_afc.delete(0, ctk.END)  # Clear the existing value
    from_ao_afc.insert(0, db_from)  # Set the new value
    from_ao_afc.configure(state="readonly")  # Set the entry back to readonly state
    
    ptr_label = ctk.CTkLabel(sub_frame2, text='PTR No.: ')
    ptr_label.pack(side = 'left', pady = 5, padx = 10)
    
    ptr_no = ctk.CTkEntry(sub_frame2, fg_color='white' , width = 155, state="readonly")
    ptr_no.pack(side = 'left', pady = 5, padx = 10)
    ptr_no.configure(state="normal")  # Set the entry to normal state
    ptr_no.delete(0, ctk.END)  # Clear the existing value
    ptr_no.insert(0, db_ptr)  # Set the new value
    ptr_no.configure(state="readonly")  # Set the entry back to readonly state
    
    sub_frame3 = ctk.CTkFrame(frame, fg_color='transparent')
    sub_frame3.pack(fill = 'x', pady = 5, padx = 10)
    
    to_label = ctk.CTkLabel(sub_frame3, text='To Accountable Officer/Agency Cluster Fund: ')
    to_label.pack(side = 'left', pady = 5, padx = 10)
    to_ao_afc = ctk.CTkEntry(sub_frame3, width = 300, fg_color='white',state="readonly")
    to_ao_afc.pack(side = 'left', pady = 5, padx = 10)
    to_ao_afc.configure(state="normal")  # Set the entry to normal state
    to_ao_afc.delete(0, ctk.END)  # Clear the existing value
    to_ao_afc.insert(0, db_to)  # Set the new value
    to_ao_afc.configure(state="readonly")  # Set the entry back to readonly state
    
    date_label = ctk.CTkLabel(sub_frame3, text='Date: ')
    date_label.pack(side = 'left', pady = 5, padx = 24)
    date = ctk.CTkEntry(sub_frame3, width = 150, fg_color='white',state="readonly")
    date.pack(side = 'left', pady = 5, padx = 10, ipady =2, ipadx = 2)
    date.pack(side = 'left', pady = 5, padx = 10)
    date.configure(state="normal")  # Set the entry to normal state
    date.delete(0, ctk.END)  # Clear the existing value
    date.insert(0, db_dt)  # Set the new value
    date.configure(state="readonly")  # Set the entry back to readonly state
    #transfer types
    frame2 = ctk.CTkFrame(tab1, corner_radius=20, fg_color='#F4F4F4')
    frame2.pack(fill = 'both', pady = 5, padx = 10)
    
    sub_frame4 = ctk.CTkFrame(frame2, fg_color='transparent')
    sub_frame4.pack(fill = 'x')
    
    transfer_type_label = ctk.CTkLabel(sub_frame4, text='Transfer Type (Select Only One): ')
    transfer_type_label.pack(side = 'left', pady = 10, padx = 10)
    
    
    transfer_var1 = ctk.StringVar()
    transfer_var2 = ctk.StringVar()
    transfer_var3 = ctk.StringVar()
    transfer_var4 = ctk.StringVar()
    transfer_var1.set("on")
    
    if db_type == 'Donation':
        transfer_var1.set("on")
        transfer_var2.set("off")
        transfer_var3.set("off")
        transfer_var4.set("off") 
        
        transfer_1 = ctk.CTkCheckBox(sub_frame4, state="disabled", text='Donation', variable=transfer_var1, onvalue="on", offvalue="off")
        transfer_1.pack(side='left', pady=5, padx=10)

        transfer_2 = ctk.CTkCheckBox(sub_frame4, state="disabled", text='Relocate', variable=transfer_var2, onvalue="on", offvalue="off")
        transfer_2.pack(side='left', pady=5, padx=10)

        transfer_3 = ctk.CTkCheckBox(sub_frame4, state="disabled", text='Reassign', variable=transfer_var3, onvalue="on", offvalue="off")
        transfer_3.pack(side='left', pady=5, padx=10)

        transfer_4 = ctk.CTkCheckBox(sub_frame4, state="disabled", text='Others (Specify): ',  variable=transfer_var4, onvalue="on", offvalue='off')
        transfer_4.pack(side='left', pady=5, padx=10)
        
    elif db_type == 'Relocate':
        transfer_var2.set("on")
        transfer_var1.set("off")
        transfer_var3.set("off")
        transfer_var4.set("off") 
         
        transfer_1 = ctk.CTkCheckBox(sub_frame4, state="disabled", text='Donation', variable=transfer_var1, onvalue="on", offvalue="off")
        transfer_1.pack(side='left', pady=5, padx=10)

        transfer_2 = ctk.CTkCheckBox(sub_frame4, state="disabled", text='Relocate', variable=transfer_var2, onvalue="on", offvalue="off")
        transfer_2.pack(side='left', pady=5, padx=10)

        transfer_3 = ctk.CTkCheckBox(sub_frame4, state="disabled", text='Reassign', variable=transfer_var3, onvalue="on", offvalue="off")
        transfer_3.pack(side='left', pady=5, padx=10)

        transfer_4 = ctk.CTkCheckBox(sub_frame4, state="disabled", text='Others (Specify): ',  variable=transfer_var4, onvalue="on", offvalue='off')
        transfer_4.pack(side='left', pady=5, padx=10)  
    elif db_type == 'Reassign':
        transfer_var3.set("on")
        transfer_var1.set("off")
        transfer_var2.set("off")
        transfer_var4.set("off")
        
        transfer_1 = ctk.CTkCheckBox(sub_frame4, state="disabled", text='Donation', variable=transfer_var1, onvalue="on", offvalue="off")
        transfer_1.pack(side='left', pady=5, padx=10)

        transfer_2 = ctk.CTkCheckBox(sub_frame4, state="disabled", text='Relocate', variable=transfer_var2, onvalue="on", offvalue="off")
        transfer_2.pack(side='left', pady=5, padx=10)

        transfer_3 = ctk.CTkCheckBox(sub_frame4, state="disabled", text='Reassign', variable=transfer_var3, onvalue="on", offvalue="off")
        transfer_3.pack(side='left', pady=5, padx=10)

        transfer_4 = ctk.CTkCheckBox(sub_frame4, state="disabled", text='Others (Specify): ',  variable=transfer_var4, onvalue="on", offvalue='off')
        transfer_4.pack(side='left', pady=5, padx=10)
    else:
        transfer_var4.set("on")
        transfer_var1.set("off")
        transfer_var2.set("off")
        transfer_var3.set("off")
        
        transfer_1 = ctk.CTkCheckBox(sub_frame4, state="disabled", text='Donation', variable=transfer_var1, onvalue="on", offvalue="off")
        transfer_1.pack(side='left', pady=5, padx=10)

        transfer_2 = ctk.CTkCheckBox(sub_frame4, state="disabled", text='Relocate', variable=transfer_var2, onvalue="on", offvalue="off")
        transfer_2.pack(side='left', pady=5, padx=10)

        transfer_3 = ctk.CTkCheckBox(sub_frame4, state="disabled", text='Reassign', variable=transfer_var3, onvalue="on", offvalue="off")
        transfer_3.pack(side='left', pady=5, padx=10)

        transfer_4 = ctk.CTkCheckBox(sub_frame4, state="disabled", text='Others (Specify): ',  variable=transfer_var4, onvalue="on", offvalue='off')
        transfer_4.pack(side='left', pady=5, padx=10)

        transfer_4_entry = ctk.CTkEntry(sub_frame4, fg_color='white',state="readonly")
        transfer_4_entry.pack(side='left', pady=5, padx=10)
        transfer_4_entry.configure(state="normal")  # Enable text editing initially
        transfer_4_entry.delete(0, ctk.END)  # Clear the existing value
        transfer_4_entry.insert(0, db_type)  # Set the new value
        transfer_4_entry.configure(state="readonly")  # Set the entry back to readonly state

    #second tab
    #items
    frame3 = ctk.CTkFrame(tab2,  fg_color='#F4F4F4')
    frame3.pack(fill = 'both', pady = 5, padx = 10)
    
    sub_frame4_3 = ctk.CTkFrame(frame3, fg_color='#F4F4F4')
    sub_frame4_3.pack(fill = 'x')
    
    
    
    item_label = ctk.CTkLabel(sub_frame4_3, text='Items')
    item_label.pack(side = 'left', pady = 5, padx = 10)

    
    
    
    
    #item table
    sub_frame4_5 = ctk.CTkFrame(tab2, fg_color='#F4F4F4')
    sub_frame4_5.pack(fill = 'x', pady = 5, padx = 10)
    
    columns = ('date_acquired', 'property_no', 'name',  'description', 'quantity', 'condition','location')
    
    item_table_report = ttk.Treeview(sub_frame4_5, columns=columns, show='headings', height=10)
    item_table_report.pack(fill=tk.BOTH)
    
    item_table_report.column('date_acquired', width = 100)
    item_table_report.column('property_no', width = 100)
    item_table_report.column('name', width = 100)
    item_table_report.column('description', width = 100)
    item_table_report.column('quantity', width = 100)
    item_table_report.column('condition', width = 100)
    item_table_report.column('location', width = 100)
    
    item_table_report.heading('date_acquired', text='Date Acquired')
    item_table_report.heading('property_no', text='Property No.')
    item_table_report.heading('name', text='Name')
    item_table_report.heading('description', text='Description')
    item_table_report.heading('quantity', text='Quantity')
    item_table_report.heading('condition', text='Condition')
    item_table_report.heading('location', text='Location')
    # Get all item IDs
    item_ids = item_table_report.get_children()

    # Delete each item
    for item_id in item_ids:
        item_table_report.delete(item_id)
        
    for item in items:
        item_table_report.insert('', 'end', values=item)
    #reasons
    frame4 = ctk.CTkFrame(tab2)
    frame4.pack(fill = 'both', pady = 5, padx = 10)
    
    sub_frame5 = ctk.CTkFrame(frame4, fg_color='transparent')
    sub_frame5.pack(fill = 'x')
    reason_label = ctk.CTkLabel(sub_frame5, text = 'Reasons for Transfer: ')
    reason_label.pack(side = 'left', pady = 5, padx = 10)
    
    
        
    entry = ctk.CTkTextbox(frame4, height=90, fg_color='white', wrap="word", state="disabled")
    entry.pack(fill='both', pady=10, padx=10)

    entry.configure(state="normal")  # Enable text editing initially
    entry.delete(1.0, ctk.END)  # Clear the existing value
    entry.insert(1.0, reason)  # Set the new value
    entry.configure(state="disabled")  # Set the entry back to readonly state

    #third tab
    #approved by
    frame5 = ctk.CTkFrame(tab3, fg_color='#F4F4F4')
    frame5.pack(fill = 'both', pady = 5, padx = 10)
    
    sub_frame6 = ctk.CTkFrame(frame5, fg_color='transparent')
    sub_frame6.pack(fill = 'x')
    
    approved_by_label = ctk.CTkLabel(sub_frame6, text = 'Approved by: ')
    approved_by_label.pack(side = 'left', pady = 3, padx = 10)
    
    sub_frame6_2 = ctk.CTkFrame(frame5, fg_color='transparent')
    sub_frame6_2.pack(fill = 'x')
    
    name_label = ctk.CTkLabel(sub_frame6_2, text = 'Printed Name: ')
    name_label.pack(side = 'left', pady = 3, padx = 10)
    approve_name = ctk.CTkEntry(sub_frame6_2, width = 300, fg_color='white', state="readonly")
    approve_name.pack(side = 'left', pady = 3, padx = 10)
    approve_name.pack(fill="both", expand=True)
    approve_name.configure(state="normal")  # Set the entry to normal state
    approve_name.delete(0, ctk.END)  # Clear the existing value
    approve_name.insert(0, db_name)  # Set the new value
    approve_name.configure(state="readonly")  # Set the entry back to readonly state
    
    
    designation_label = ctk.CTkLabel(sub_frame6_2, text = 'Designation: ')
    designation_label.pack(side = 'left', pady = 3, padx = 10)
    approve_designation = ctk.CTkEntry(sub_frame6_2, width = 200, fg_color='white', state="readonly")
    approve_designation.pack(side = 'left', pady = 3, padx = 10)
    approve_designation.configure(state="normal")  # Set the entry to normal state
    approve_designation.delete(0, ctk.END)  # Clear the existing value
    approve_designation.insert(0, db_des)  # Set the new value
    approve_designation.configure(state="readonly")  # Set the entry back to readonly state
    
    
    approve_date_label = ctk.CTkLabel(sub_frame6_2, text = 'Date: ')
    approve_date_label.pack(side = 'left', pady = 3, padx = 10)
    approve_date = ctk.CTkEntry(sub_frame6_2, width = 150, fg_color='white',state="readonly")
    approve_date.pack(side = 'left', pady = 3, padx = 10)
    approve_date.configure(state="normal")  # Set the entry to normal state
    approve_date.delete(0, ctk.END)  # Clear the existing value
    approve_date.insert(0, db_adt)  # Set the new value
    approve_date.configure(state="readonly")  # Set the entry back to readonly state
    
    #Released/issued by
    frame6 = ctk.CTkFrame(tab3, fg_color='#F4F4F4')
    frame6.pack(fill = 'both', pady = 5, padx = 10) 
    
    sub_frame7 = ctk.CTkFrame(frame6, fg_color='transparent')
    sub_frame7.pack(fill = 'x')
    
    released_by_label = ctk.CTkLabel(sub_frame7, text = 'Released/Issued by: ')
    released_by_label.pack(side = 'left', pady = 3, padx = 10)
    
    sub_frame7_2 = ctk.CTkFrame(frame6, fg_color='transparent')
    sub_frame7_2.pack(fill = 'x')
    
    name_label = ctk.CTkLabel(sub_frame7_2, text = 'Printed Name: ')
    name_label.pack(side = 'left', pady = 3, padx = 10)
    release_name = ctk.CTkEntry(sub_frame7_2, width = 300, fg_color='white', state="readonly")
    release_name.pack(side = 'left', pady = 3, padx = 10)
    release_name.configure(state="normal")  # Set the entry to normal state
    release_name.delete(0, ctk.END)  # Clear the existing value
    release_name.insert(0, db_rname)  # Set the new value
    release_name.configure(state="readonly")  # Set the entry back to readonly state
    
    
    designation_label = ctk.CTkLabel(sub_frame7_2, text = 'Designation: ')
    designation_label.pack(side = 'left', pady = 3, padx = 10)
    release_designation = ctk.CTkEntry(sub_frame7_2, width = 200, fg_color='white', state="readonly")
    release_designation.pack(side = 'left', pady = 3, padx = 10)
    release_designation.configure(state="normal")  # Set the entry to normal state
    release_designation.delete(0, ctk.END)  # Clear the existing value
    release_designation.insert(0, db_rdes)  # Set the new value
    release_designation.configure(state="readonly")  # Set the entry back to readonly state
    
    
    release_date_label = ctk.CTkLabel(sub_frame7_2, text = 'Date: ')
    release_date_label.pack(side = 'left', pady = 3, padx = 10)
    release_date = ctk.CTkEntry(sub_frame7_2,  width = 150, fg_color='white',state="readonly" )
    release_date.pack(side = 'left', pady = 3, padx = 10)
    release_date.configure(state="normal")  # Set the entry to normal state
    release_date.delete(0, ctk.END)  # Clear the existing value
    release_date.insert(0, db_rdt)  # Set the new value
    release_date.configure(state="readonly")  # Set the entry back to readonly state
    #receive by
    frame7 = ctk.CTkFrame(tab3, fg_color='#F4F4F4')
    frame7.pack(fill = 'both', pady = 5, padx = 10) 
    
    sub_frame8 = ctk.CTkFrame(frame7, fg_color='transparent')
    sub_frame8.pack(fill = 'x')
    
    received_by_label = ctk.CTkLabel(sub_frame8, text = 'Received by: ')
    received_by_label.pack(side = 'left', pady = 5, padx = 10)
    
    sub_frame8_2 = ctk.CTkFrame(frame7, fg_color='transparent')
    sub_frame8_2.pack(fill = 'x')
    
    name_label = ctk.CTkLabel(sub_frame8_2, text = 'Printed Name: ')
    name_label.pack(side = 'left', pady = 2, padx = 10)
    receive_name = ctk.CTkEntry(sub_frame8_2, width = 300, fg_color='white', state="readonly")
    receive_name.pack(side = 'left', pady = 2, padx = 10)
    receive_name.configure(state="normal")  # Set the entry to normal state
    receive_name.delete(0, ctk.END)  # Clear the existing value
    receive_name.insert(0, db_riname)  # Set the new value
    receive_name.configure(state="readonly")  # Set the entry back to readonly state
    
    designation_label = ctk.CTkLabel(sub_frame8_2, text = 'Designation: ')
    designation_label.pack(side = 'left', pady = 2, padx = 10)
    receive_designation = ctk.CTkEntry(sub_frame8_2, width= 200, fg_color='white', state="readonly")
    receive_designation.pack(side = 'left', pady = 2, padx = 10)
    receive_designation.configure(state="normal")  # Set the entry to normal state
    receive_designation.delete(0, ctk.END)  # Clear the existing value
    receive_designation.insert(0, db_rides)  # Set the new value
    receive_designation.configure(state="readonly")  # Set the entry back to readonly state
    
    
    receive_date_label = ctk.CTkLabel(sub_frame8_2, text = 'Date: ')
    receive_date_label.pack(side = 'left', pady = 5, padx = 10)
    receive_date = ctk.CTkEntry(sub_frame8_2,  width = 150, fg_color='white',state="readonly" )
    receive_date.pack(side = 'left', pady = 2, padx = 10)
    receive_date.configure(state="normal")  # Set the entry to normal state
    receive_date.delete(0, ctk.END)  # Clear the existing value
    receive_date.insert(0, db_ridt)  # Set the new value
    receive_date.configure(state="readonly")  # Set the entry back to readonly state
    
    if db_remarks == 'REJECTED':
        #reject report
        c_reject = ctk.CTkButton(main, command=lambda val='DELETE': approval_report(val), text = 'DELETE', font= ctk.CTkFont('Arial', size = 15, weight = "bold"), fg_color='#bf0e06', hover_color='#ec4c47',  text_color='white', corner_radius = 30)
        c_reject.pack(side = 'right', pady = 5, padx = (0, 20))
    else:
        #approve report
        c_approve = ctk.CTkButton(main, command=lambda val='APPROVED': approval_report(val), text = 'GENERATE', font= ctk.CTkFont('Arial', size = 15, weight = "bold"), fg_color='#187bcd', hover_color='#2a9df4', text_color='white', corner_radius = 30)
        c_approve.pack(side = 'right', pady = 5, padx = (0, 20))
    
    
            
    def callback():
        from tkinter import messagebox
        from resources.py_client.client_translist import requestlist_window2
        try:
            report_form.withdraw()
            requestlist_window2()
            
        except Exception as e:
            if messagebox.askok("ERROR", e ):
                print(e)
                

    report_form.protocol("WM_DELETE_WINDOW", callback)
    report_form.mainloop()

    
            
                
#showptr_window()
