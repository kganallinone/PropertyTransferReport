
from tkinter import *
import tkinter as tk
import customtkinter as ctk 
from PIL import Image 
import pyrebase 
from tkinter import messagebox 
import openpyxl

def login_window():
    
    from resources.py_admin.page1_dashboard import dashboard_window
    from resources.py_client.client_dashboard import dashboard_window_client
    from resources.py_home.register import register_window
    
    ctk.set_appearance_mode("light")  # Modes: system (default), light, dark
    
    #<----------------FIREBASE CONFIGURATION--------------------------------->#
    config = {
        "apiKey": "AIzaSyCyCtXg05ff-tocdUaxjBFfa550TLfKZQ4",
        "authDomain": "puplopez-ptp.firebaseapp.com",
        "projectId": "puplopez-ptp",
        "databaseURL": "https://puplopez-ptp-default-rtdb.firebaseio.com/",
        "storageBucket": "puplopez-ptp.appspot.com",
        "messagingSenderId": "638312293451",
        "appId": "1:638312293451:web:79e9405d11db0496f599ce",
        "measurementId": "G-9Z3SPKFR9V"
    }
    #<----------------FIREBASE CONNECT--------------------------------->#
    firebase = pyrebase.initialize_app(config)


    # Get a reference to the auth service
    auth = firebase.auth()
    db = firebase.database()
    #<----------------FIREBASE END--------------------------------->#

   
    login = tk.Toplevel()
    login.title("Property Transfer | login")
    login.geometry("{0}x{1}+200+50".format(1000, 600))
    login.resizable(False, False)
    login.configure(background='white')
    open_img = PhotoImage(file = 'resources\images\DEFAULT\PTR.png')   
    login.iconphoto(False, open_img)   

    #------------------- lOGIN FUNCTION ------------------------------------#
    
    def button_function():
       
        email = login_email.get()
        pw = login_pw.get()
        
        substring = "@"
        substring2 = ".com"

        if email != None and substring in email and substring2 in email:
            
            users_db = db.child("user_accounts").order_by_child("EMAIL").equal_to(email).get()
        
            for dt in users_db.each():
                userstype = dt.val()['TYPE']
                useremail = dt.val()['EMAIL']
                userpw = dt.val()['PW']
            
            try:
                auth.sign_in_with_email_and_password(email, pw)
                
                
                if userstype == 'Admin':
                    workbook = openpyxl.Workbook()
                    ws1 = workbook.active
                    ws1.title = 'User email'
                    ws2 = workbook.create_sheet('Report')
                    ws1['A1'] = email
                    ws2['A1'] = ''
                    workbook.save('./resources/user/data/UserLog.xlsx')
                    
                    login.withdraw()
                    dashboard_window()
                    
                    
                else:
                    workbook = openpyxl.Workbook()
                    ws1 = workbook.active
                    ws1.title = 'User email'
                    ws2 = workbook.create_sheet('Report')
                    ws1['A1'] = email
                    ws2['A1'] = ''
                    workbook.save('./resources/user/data/UserLog.xlsx')
                    
                    
                    login.withdraw()
                    dashboard_window_client()
                    

                
            except:
                
                if email == useremail and pw == userpw:
                
                    if userstype == 'Admin':
                        workbook = openpyxl.Workbook()
                        ws1 = workbook.active
                        ws1.title = 'User email'
                        ws2 = workbook.create_sheet('Report')
                        ws1['A1'] = email
                        ws2['A1'] = ''
                        workbook.save("./resources/user/data/UserLog.xlsx")
                    
                        
                        login.withdraw()
                        dashboard_window()
                        
                        
                    else:
                        workbook = openpyxl.Workbook()
                        ws1 = workbook.active
                        ws1.title = 'User email'
                        ws2 = workbook.create_sheet('Report')
                        ws1['A1'] = email
                        ws2['A1'] = ''
                        workbook.save('./resources/user/data/UserLog.xlsx')
                        
                        login.withdraw()
                        dashboard_window_client()
                    
                else:
                    login_pw.delete(0,len(pw))
                    answer = messagebox.askyesno("Question","Invalid user! Register first!")
                    print(answer)
        
        elif len(email) == 5 or len(email) == 6:
            
            users_db = db.child("user_accounts").order_by_child("ID").equal_to(email).get()
        
            for dt in users_db.each():
                userstype = dt.val()['TYPE']
                usersemail = dt.val()['EMAIL']
            
            try:
                auth.sign_in_with_email_and_password(usersemail, pw)
                
                
                if userstype == 'Admin':
                    workbook = openpyxl.Workbook()
                    ws1 = workbook.active
                    ws1.title = 'User email'
                    ws2 = workbook.create_sheet('Report')
                    ws1['A1'] = email
                    ws2['A1'] = ''
                    workbook.save('./resources/user/data/UserLog.xlsx')
                    
                    login.withdraw()
                    dashboard_window()
                    
                    
                else:
                    workbook = openpyxl.Workbook()
                    ws1 = workbook.active
                    ws1.title = 'User email'
                    ws2 = workbook.create_sheet('Report')
                    ws1['A1'] = email
                    ws2['A1'] = ''
                    workbook.save('./resources/user/data/UserLog.xlsx')
                    
                    login.withdraw()
                    dashboard_window_client()
                    
                    
                
                
            except:
                login_pw.delete(0,len(pw))
                answer = messagebox.askyesno("Question","Invalid user! Register first!")
                print(answer)    
            
        else:
            login_email.delete(0,len(email))
            login_pw.delete(0,len(pw))
            answer = messagebox.askyesno("Question","Invalid user! Register first!")
            print(answer)
        
            
        
        
            
        

    
    #------------------- DESIGN FUNCTION ------------------------------------#



    #------------------- uI DESIGN------------------------------------#
    '''
    login_openimg = Image.open(r'resources\images\signup_bg.png')
    #resize_image = openimage1.resize(( login.winfo_width(),login.winfo_height()))
    login_resize_image = login_openimg.resize((login.winfo_screenwidth(),login.winfo_screenheight()))
    login_img1=ImageTk.PhotoImage(login_resize_image)
    login_label= Label(login, image=login_img1)
    login_label.image = login_img1
    login_label.pack()
    '''
    
    login_form_bg_image = ctk.CTkImage(Image.open('resources\images\HOME\signup_bg.png'), size=(login.winfo_screenwidth(), login.winfo_screenheight()))
    login_form_bg = ctk.CTkLabel(login, image=login_form_bg_image, text=None, anchor='center')
    login_form_bg.place(x=0, y=0, relwidth=1, relheight=1)
    
    #frame for register form
    login_form_frame = ctk.CTkFrame(login, fg_color='transparent')
    login_form_frame.pack(side = 'right', pady = 15, padx = 75)
   

    
    login_label1=ctk.CTkLabel( 
        login_form_frame, 
        text="LOG-IN",
        text_color='#3B3B3B',
        font = ctk.CTkFont (
        'Arial', 
        weight="bold",
        size= 40
        ), 
        fg_color='#FFFFFF'
        )
    login_label1.pack(pady = (15, 15), padx = 10)

    login_email=ctk.CTkEntry(
        login_form_frame, 
        width=320, 
        height=40,
        corner_radius= 30,
        border_color='#C5C5C5',
        placeholder_text='Email/ID',
        placeholder_text_color='#525252',
        fg_color='#C5C5C5',
        text_color='#3B3B3B'
        )

    login_email.pack(pady = 10, padx = 10)


    login_pw=ctk.CTkEntry(
        login_form_frame, 
        width=320, 
        height=40,
        corner_radius= 30,
        border_color='#C5C5C5',
        placeholder_text='Password',
        placeholder_text_color='#525252',
        fg_color='#C5C5C5',
        text_color='#3B3B3B',
        show= '●'
        )
    login_pw.pack(pady = 10, padx = 10)

    login_label2=ctk.CTkLabel(
        login_form_frame, 
        text="Forget Password?",
        text_color='#525252',
        font = ctk.CTkFont (
        'Arial', 
        weight="bold",
        size= 12
        ), 
        fg_color='#FFFFFF'
        )

    login_label2.pack(pady = 10, padx = 10)

    #Create custom button
    login_bttn1 = ctk.CTkButton(
        login_form_frame, 
        command=lambda: button_function(), 
        width=150, 
        height=40, 
        compound="left", 
        corner_radius= 30, 
        fg_color='maroon', 
        text_color='white', 
        hover_color='#4F0000', 
        font = ctk.CTkFont (
            'Arial', 
            weight="bold"
            ), 
        text="LOG-IN" 
        )

    login_bttn1.pack(pady = 10, padx = 10)
    
    #for register
    goto_reg_frame = ctk.CTkFrame(login_form_frame, fg_color='transparent')
    goto_reg_frame.pack(pady = (15,0))
    
    login_label3=ctk.CTkLabel(
        goto_reg_frame , 
        text="Don't have an account yet?",
        text_color='#525252',
        font = ctk.CTkFont (
        'Arial', 
        weight="bold",
        size= 12
        ), 
        fg_color='#FFFFFF'
        )

    login_label3.pack(side = 'left')
    
    def register_open():
        login.withdraw()
        register_window()
        
    login_label4=ctk.CTkLabel(
        goto_reg_frame ,
        text="Register.",
        text_color='#0000FF',
        font = ctk.CTkFont (
        'Arial', 
        weight="bold",
        size= 12
        ), 
        fg_color='#FFFFFF'
        )

    login_label4.pack(side = 'left', padx = 3)
    login_label4.bind(
        "<Button-1>", 
        lambda e:
        register_open()
        )
    def callback():
        try:
            import subprocess
            progs = str(subprocess.check_output('tasklist'))
            process_name = "PropertyTransfer.exe"
            if process_name in progs:
                if messagebox.askokcancel("Quit", "Do you really wish to quit?"):
                    subprocess.call("TASKKILL /F /IM PropertyTransfer.exe", shell=True)
                    login.destroy()
            else:
                if messagebox.askokcancel("Quit", "Do you really wish to quit?"):
                    login.destroy()
        
        except Exception as e:
            if messagebox.askok("ERROR", e ):
                print(e)

    login.protocol("WM_DELETE_WINDOW", callback)
    login.mainloop()

    
#login_window()